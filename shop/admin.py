import mimetypes
import shutil
from collections import OrderedDict
import re
import uuid
from io import BytesIO
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen
from zipfile import ZIP_DEFLATED, ZipFile

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.template.response import TemplateResponse
from django.urls import path, reverse
from django.utils import timezone
from django.utils.html import escape, format_html
from django.utils.safestring import mark_safe
from tinymce.widgets import TinyMCE

from shop.file_utils import infer_file_kind
from shop.models import (
    Brand,
    Characteristic,
    City,
    Group,
    HtmlContent,
    Inquiry,
    Agent,
    News,
    NewsAttachment,
    MediaLibrary,
    Product,
    ProductCertificate,
    ProductGalleryItem,
    ProductCharacteristic,
    ProductMedia,
    OrderEmailRecipient,
    OrderEmailSettings,
    PUBLISH_STATUS_DRAFT,
    PUBLISH_STATUS_PUBLISHED,
    PublicOrder,
    PublicOrderItem,
    Sert,
    Slider,
    ContactInfo,
)
from shop.services.order_email import build_notification_preview_html


CYRILLIC_TO_LATIN = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}


def transliterate_slug(value: str) -> str:
    normalized = str(value or "").strip().lower()
    transliterated = "".join(CYRILLIC_TO_LATIN.get(char, char) for char in normalized)
    transliterated = transliterated.replace("&", " and ")
    transliterated = re.sub(r"[^a-z0-9]+", "-", transliterated)
    transliterated = re.sub(r"-{2,}", "-", transliterated).strip("-")
    return transliterated or f"item-{uuid.uuid4().hex[:8]}"


def parse_bool(value):
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().lower()
    if normalized in {"да", "true", "1", "yes", "y"}:
        return True
    if normalized in {"нет", "false", "0", "no", "n"}:
        return False
    return bool(normalized)


def parse_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValidationError(f"Cannot parse decimal value '{value}'") from exc


def split_media_urls(value):
    if not value:
        return None
    if isinstance(value, list):
        urls = [str(item).strip() for item in value if str(item).strip()]
        return urls or None
    urls = [item.strip() for item in re.split(r"[\n;,]+", str(value)) if item.strip()]
    return urls or None


def is_probable_url(value: str | None) -> bool:
    if not value:
        return False
    parsed = urlparse(str(value).strip())
    return parsed.scheme in {"http", "https"} or str(value).startswith(str(settings.MEDIA_URL))


def normalize_local_file_path(value: str | None) -> Path | None:
    text = str(value or "").strip().strip('"')
    if not text:
        return None
    parsed = urlparse(text)
    if parsed.scheme == "file":
        text = parsed.path
        if re.match(r"^/[a-zA-Z]:/", text):
            text = text[1:]
    if is_probable_url(text):
        return None
    path = Path(text).expanduser()
    if path.exists() and path.is_file():
        return path
    return None


def save_local_file_path(source_path: Path, folder_name: str) -> dict:
    media_root = Path(settings.MEDIA_ROOT)
    target_dir = media_root / "admin_uploads" / folder_name
    target_dir.mkdir(parents=True, exist_ok=True)

    suffix = source_path.suffix
    filename = f"{uuid.uuid4().hex}{suffix}"
    storage_path = target_dir / filename
    shutil.copyfile(source_path, storage_path)

    relative_path = storage_path.relative_to(media_root).as_posix()
    mime_type = mimetypes.guess_type(source_path.name)[0] or "application/octet-stream"

    return {
        "storage_path": str(storage_path),
        "url": f"{settings.MEDIA_URL}{relative_path}",
        "mime_type": mime_type,
        "size_bytes": storage_path.stat().st_size,
        "title": source_path.name,
    }


def filename_from_remote_response(url: str, response) -> str:
    content_disposition = response.headers.get("Content-Disposition", "")
    match = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', content_disposition, flags=re.IGNORECASE)
    if match:
        filename = unquote(match.group(1)).strip()
        if filename:
            return Path(filename).name
    parsed_name = Path(unquote(urlparse(url).path)).name
    return parsed_name or f"remote-{uuid.uuid4().hex[:8]}"


def save_remote_file_url(url: str, folder_name: str) -> dict:
    media_root = Path(settings.MEDIA_ROOT)
    target_dir = media_root / "admin_uploads" / folder_name
    target_dir.mkdir(parents=True, exist_ok=True)

    request = Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (compatible; NovotechMediaImporter/1.0)",
        },
    )
    with urlopen(request, timeout=REMOTE_IMPORT_TIMEOUT_SECONDS) as response:
        content_length = response.headers.get("Content-Length")
        if content_length and int(content_length) > REMOTE_IMPORT_MAX_BYTES:
            raise ValidationError(f"Remote file is larger than {REMOTE_IMPORT_MAX_BYTES // 1024 // 1024} MB: {url}")

        original_name = filename_from_remote_response(url, response)
        suffix = Path(original_name).suffix
        filename = f"{uuid.uuid4().hex}{suffix}"
        storage_path = target_dir / filename
        downloaded = 0

        with storage_path.open("wb") as destination:
            while True:
                chunk = response.read(1024 * 1024)
                if not chunk:
                    break
                downloaded += len(chunk)
                if downloaded > REMOTE_IMPORT_MAX_BYTES:
                    storage_path.unlink(missing_ok=True)
                    raise ValidationError(f"Remote file is larger than {REMOTE_IMPORT_MAX_BYTES // 1024 // 1024} MB: {url}")
                destination.write(chunk)

    relative_path = storage_path.relative_to(media_root).as_posix()
    mime_type = mimetypes.guess_type(original_name)[0] or mimetypes.guess_type(str(storage_path))[0] or "application/octet-stream"

    return {
        "storage_path": str(storage_path),
        "url": f"{settings.MEDIA_URL}{relative_path}",
        "mime_type": mime_type,
        "size_bytes": storage_path.stat().st_size,
        "title": original_name,
    }


def resolve_import_file_reference(value, folder_name: str):
    text = str(value or "").strip()
    if not text:
        return None, None

    local_path = normalize_local_file_path(text)
    if local_path:
        return save_local_file_path(local_path, folder_name), None

    parsed = urlparse(text)
    if parsed.scheme in {"http", "https"}:
        try:
            return save_remote_file_url(text, folder_name), None
        except Exception as exc:
            return None, f"Remote file could not be downloaded: {text}. {exc}"

    if is_probable_url(text):
        mime_type = guess_media_mime_type(text)
        storage_path = resolve_media_storage_path(url=text) if is_local_media_url(text) else ""
        size_bytes = Path(storage_path).stat().st_size if storage_path and Path(storage_path).exists() else 0
        return {
            "storage_path": storage_path or "",
            "url": text,
            "mime_type": mime_type,
            "size_bytes": size_bytes,
            "title": Path(urlparse(text).path).name or text,
        }, None

    return None, f"File path or URL was not found: {text}"


def split_title_values(value):
    items = split_media_urls(value)
    return items or []


def infer_data_type(value) -> str:
    if value in (None, ""):
        return "text"
    if isinstance(value, (int, float, Decimal)):
        return "number"
    text = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        Decimal(text)
        return "number"
    except InvalidOperation:
        return "text"


def characteristic_name_from_header(header: str) -> str:
    return header[len("char_"):].replace("_", " ").strip()


def save_admin_upload(upload, folder_name: str) -> dict:
    media_root = Path(settings.MEDIA_ROOT)
    target_dir = media_root / "admin_uploads" / folder_name
    target_dir.mkdir(parents=True, exist_ok=True)

    suffix = Path(upload.name).suffix
    filename = f"{uuid.uuid4().hex}{suffix}"
    storage_path = target_dir / filename

    with storage_path.open("wb+") as destination:
        for chunk in upload.chunks():
            destination.write(chunk)

    relative_path = storage_path.relative_to(media_root).as_posix()
    url = f"{settings.MEDIA_URL}{relative_path}"
    mime_type = upload.content_type or mimetypes.guess_type(upload.name)[0] or "application/octet-stream"

    return {
        "storage_path": str(storage_path),
        "url": url,
        "mime_type": mime_type,
        "size_bytes": upload.size,
    }


def validate_new_file_upload(form, upload_field_name: str):
    if not getattr(form, "cleaned_data", None):
        return
    if not form.has_changed():
        return
    if form.instance.pk:
        return
    if form.cleaned_data.get(upload_field_name):
        return
    raise ValidationError({upload_field_name: "Upload a file before saving this item."})


def mark_generated_file_fields_optional(form, extra_fields=None):
    optional_fields = ["storage_path", "url", "mime_type", "size_bytes"]
    if extra_fields:
        optional_fields.extend(extra_fields)
    for field_name in optional_fields:
        if field_name in form.fields:
            form.fields[field_name].required = False


def extract_media_url(value):
    if not value:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, dict):
        for key in ("url", "src", "path"):
            if value.get(key):
                return value[key]
        return None
    if isinstance(value, list):
        for item in value:
            url = extract_media_url(item)
            if url:
                return url
    return None


def iter_json_media_items(value):
    if value in (None, "", []):
        return
    items = value if isinstance(value, list) else [value]
    for item in items:
        if isinstance(item, str):
            url = item.strip()
            if url:
                yield {"url": url, "storage_path": None, "title": None}
            continue
        if isinstance(item, dict):
            url = item.get("url") or item.get("src") or item.get("path")
            storage_path = item.get("storage_path")
            title = item.get("title") or item.get("name") or item.get("alt")
            if url or storage_path:
                yield {"url": url, "storage_path": storage_path, "title": title}


def guess_media_mime_type(url: str | None, fallback: str | None = None) -> str:
    if fallback:
        return fallback
    guessed = mimetypes.guess_type(urlparse(str(url or "")).path)[0]
    return guessed or "application/octet-stream"


def is_local_media_url(url: str | None) -> bool:
    if not url:
        return False
    parsed = urlparse(str(url))
    return not parsed.scheme and str(url).startswith(str(settings.MEDIA_URL))


def resolve_media_storage_path(url: str | None = None, storage_path: str | None = None) -> str | None:
    if storage_path:
        path = Path(storage_path)
        if not path.is_absolute():
            path = Path(settings.MEDIA_ROOT) / path
        return str(path)
    if is_local_media_url(url):
        relative = str(url)[len(str(settings.MEDIA_URL)):].lstrip("/")
        return str(Path(settings.MEDIA_ROOT) / Path(relative))
    return None


def build_media_asset_key(url: str | None = None, storage_path: str | None = None) -> str | None:
    resolved_storage_path = resolve_media_storage_path(url=url, storage_path=storage_path)
    if resolved_storage_path:
        return f"path::{resolved_storage_path}"
    if url:
        return f"url::{url}"
    return None


def build_admin_change_url(instance) -> str | None:
    try:
        return reverse(f"admin:{instance._meta.app_label}_{instance._meta.model_name}_change", args=[instance.pk])
    except Exception:
        return None


def collect_media_library_assets(search_query: str = "", usage_filter: str = "") -> list[dict]:
    assets: dict[str, dict] = {}
    query = str(search_query or "").strip().lower()
    usage_filter = str(usage_filter or "").strip()

    def register_asset(*, url, storage_path=None, title=None, source_label, owner_label, owner_admin_url=None, mime_type=None, size_bytes=None, kind=None):
        asset_key = build_media_asset_key(url=url, storage_path=storage_path)
        if not asset_key:
            return
        resolved_storage_path = resolve_media_storage_path(url=url, storage_path=storage_path)
        detected_mime_type = guess_media_mime_type(url, mime_type)
        detected_kind = kind or infer_file_kind(detected_mime_type)
        title = title or Path(urlparse(str(url or "")).path).name or owner_label
        search_blob = " ".join(filter(None, [title, source_label, owner_label, url, resolved_storage_path, detected_mime_type])).lower()
        if usage_filter and source_label != usage_filter:
            return
        if query and query not in search_blob:
            return

        asset = assets.setdefault(
            asset_key,
            {
                "asset_key": asset_key,
                "url": url,
                "storage_path": resolved_storage_path,
                "mime_type": detected_mime_type,
                "size_bytes": size_bytes,
                "kind": detected_kind,
                "title": title,
                "usages": [],
            },
        )
        if not asset.get("url") and url:
            asset["url"] = url
        if not asset.get("storage_path") and resolved_storage_path:
            asset["storage_path"] = resolved_storage_path
        if not asset.get("size_bytes") and size_bytes:
            asset["size_bytes"] = size_bytes
        if not asset.get("mime_type") and detected_mime_type:
            asset["mime_type"] = detected_mime_type
        if not asset.get("title") and title:
            asset["title"] = title
        if not asset.get("kind") and detected_kind:
            asset["kind"] = detected_kind
        asset["usages"].append(
            {
                "source_label": source_label,
                "owner_label": owner_label,
                "owner_admin_url": owner_admin_url,
            }
        )

    for brand in Brand.objects.exclude(media__isnull=True).exclude(media__exact="").order_by("name", "id"):
        register_asset(
            url=brand.media,
            source_label="Brand image",
            owner_label=brand.name,
            owner_admin_url=build_admin_change_url(brand),
            title=brand.name,
        )

    for group in Group.objects.exclude(media__isnull=True).exclude(media__exact="").order_by("name", "id"):
        register_asset(
            url=group.media,
            source_label="Group image",
            owner_label=group.name,
            owner_admin_url=build_admin_change_url(group),
            title=group.name,
        )

    for slider in Slider.objects.exclude(image__isnull=True).exclude(image__exact="").order_by("sort_order", "id"):
        register_asset(
            url=slider.image,
            source_label="Slider image",
            owner_label=slider.title,
            owner_admin_url=build_admin_change_url(slider),
            title=slider.title,
        )

    for product in Product.objects.exclude(media__isnull=True).order_by("name", "id"):
        for item in iter_json_media_items(product.media):
            register_asset(
                url=item["url"],
                storage_path=item["storage_path"],
                source_label="Product media field",
                owner_label=f"{product.name} ({product.sku})",
                owner_admin_url=build_admin_change_url(product),
                title=item["title"] or product.name,
            )

    for news in News.objects.exclude(media__isnull=True).order_by("title", "id"):
        for item in iter_json_media_items(news.media):
            register_asset(
                url=item["url"],
                storage_path=item["storage_path"],
                source_label="News media field",
                owner_label=news.title,
                owner_admin_url=build_admin_change_url(news),
                title=item["title"] or news.title,
            )

    for media in ProductMedia.objects.select_related("product").order_by("product__name", "sort_order", "id"):
        register_asset(
            url=media.url,
            storage_path=media.storage_path,
            source_label="Product media",
            owner_label=f"{media.product.name} ({media.product.sku})",
            owner_admin_url=build_admin_change_url(media.product),
            mime_type=media.mime_type,
            size_bytes=media.size_bytes,
            kind=media.media_kind,
            title=media.alt_text or media.product.name,
        )

    for item in ProductGalleryItem.objects.select_related("product").order_by("product__name", "sort_order", "id"):
        register_asset(
            url=item.url,
            storage_path=item.storage_path,
            source_label="Product gallery",
            owner_label=f"{item.product.name} ({item.product.sku})",
            owner_admin_url=build_admin_change_url(item.product),
            mime_type=item.mime_type,
            size_bytes=item.size_bytes,
            kind=item.file_kind,
            title=item.title or item.product.name,
        )

    for certificate in ProductCertificate.objects.select_related("product").order_by("product__name", "sort_order", "id"):
        register_asset(
            url=certificate.url,
            storage_path=certificate.storage_path,
            source_label="Product certificate",
            owner_label=f"{certificate.product.name} ({certificate.product.sku})",
            owner_admin_url=build_admin_change_url(certificate.product),
            mime_type=certificate.mime_type,
            size_bytes=certificate.size_bytes,
            kind="document",
            title=certificate.title,
        )

    for attachment in NewsAttachment.objects.select_related("news").order_by("news__title", "sort_order", "id"):
        register_asset(
            url=attachment.url,
            storage_path=attachment.storage_path,
            source_label="News attachment",
            owner_label=attachment.news.title,
            owner_admin_url=build_admin_change_url(attachment.news),
            mime_type=attachment.mime_type,
            size_bytes=attachment.size_bytes,
            kind=infer_file_kind(attachment.mime_type),
            title=attachment.title,
        )

    for sert in Sert.objects.order_by("sort_order", "id"):
        register_asset(
            url=sert.url,
            storage_path=sert.storage_path,
            source_label="Sert file",
            owner_label=sert.title,
            owner_admin_url=build_admin_change_url(sert),
            mime_type=sert.mime_type,
            size_bytes=sert.size_bytes,
            kind="document",
            title=sert.title,
        )

    prepared_assets = []
    for asset in assets.values():
        asset["usage_count"] = len(asset["usages"])
        asset["is_local"] = bool(asset.get("storage_path"))
        asset["file_exists"] = bool(asset.get("storage_path") and Path(asset["storage_path"]).exists())
        asset["preview_url"] = asset["url"] if asset.get("kind") == "image" else None
        prepared_assets.append(asset)

    return sorted(prepared_assets, key=lambda item: (item["title"].lower(), item.get("url") or "", item["asset_key"]))


def media_item_matches(value, asset_url: str | None, asset_storage_path: str | None) -> bool:
    item_url = None
    item_storage_path = None
    if isinstance(value, str):
        item_url = value.strip()
    elif isinstance(value, dict):
        item_url = value.get("url") or value.get("src") or value.get("path")
        item_storage_path = value.get("storage_path")
    resolved_storage_path = resolve_media_storage_path(url=item_url, storage_path=item_storage_path)
    return bool(
        (asset_url and item_url == asset_url)
        or (asset_storage_path and resolved_storage_path == asset_storage_path)
    )


def strip_asset_from_json_media(value, asset_url: str | None, asset_storage_path: str | None):
    if value in (None, ""):
        return value, False
    if isinstance(value, list):
        filtered = [item for item in value if not media_item_matches(item, asset_url, asset_storage_path)]
        changed = len(filtered) != len(value)
        return (filtered or None), changed
    if media_item_matches(value, asset_url, asset_storage_path):
        return None, True
    return value, False


def build_media_match_q(asset_url: str | None, asset_storage_path: str | None):
    conditions = Q(pk__in=[])
    if asset_url:
        conditions |= Q(url=asset_url)
    if asset_storage_path:
        conditions |= Q(storage_path=asset_storage_path)
    return conditions


def delete_media_asset(asset_url: str | None, asset_storage_path: str | None) -> dict:
    resolved_storage_path = resolve_media_storage_path(url=asset_url, storage_path=asset_storage_path)
    q = build_media_match_q(asset_url, resolved_storage_path)
    affected_product_ids = set()
    deleted_rows = {}

    for model, label in (
        (ProductMedia, "product_media"),
        (ProductGalleryItem, "product_gallery"),
        (ProductCertificate, "product_certificates"),
        (NewsAttachment, "news_attachments"),
        (Sert, "serts"),
    ):
        queryset = model.objects.filter(q)
        if model is ProductMedia:
            affected_product_ids.update(queryset.values_list("product_id", flat=True))
        count = queryset.count()
        if count:
            queryset.delete()
        deleted_rows[label] = count

    cleared_fields = {
        "brands": Brand.objects.filter(media=asset_url).update(media=None) if asset_url else 0,
        "groups": Group.objects.filter(media=asset_url).update(media=None) if asset_url else 0,
        "sliders": Slider.objects.filter(image=asset_url).update(image="") if asset_url else 0,
    }

    updated_products = 0
    for product in Product.objects.exclude(pk__in=affected_product_ids).exclude(media__isnull=True):
        updated_media, changed = strip_asset_from_json_media(product.media, asset_url, resolved_storage_path)
        if changed:
            product.media = updated_media
            product.save(update_fields=["media"])
            updated_products += 1

    updated_news = 0
    for news in News.objects.exclude(media__isnull=True):
        updated_media, changed = strip_asset_from_json_media(news.media, asset_url, resolved_storage_path)
        if changed:
            news.media = updated_media
            news.save(update_fields=["media"])
            updated_news += 1

    for product_id in affected_product_ids:
        product = Product.objects.filter(pk=product_id).first()
        if product:
            ensure_single_primary(product)
            sync_product_media(product)

    file_deleted = False
    if resolved_storage_path:
        file_path = Path(resolved_storage_path)
        if file_path.exists() and file_path.is_file():
            file_path.unlink()
            file_deleted = True

    return {
        "deleted_rows": deleted_rows,
        "cleared_fields": cleared_fields,
        "updated_products": updated_products,
        "updated_news": updated_news,
        "file_deleted": file_deleted,
        "affected_product_count": len(affected_product_ids),
    }


TABLE_BLOCK_RE = re.compile(r"<table\b(?P<attrs>[^>]*)>(?P<content>.*?)</table>", flags=re.IGNORECASE | re.DOTALL)
ATTR_RE = re.compile(r'([a-zA-Z_:][-a-zA-Z0-9_:.]*)="([^"]*)"')
TR_BLOCK_RE = re.compile(r"<tr\b(?P<attrs>[^>]*)>(?P<content>.*?)</tr>", flags=re.IGNORECASE | re.DOTALL)
CELL_OPEN_RE = re.compile(r"<(?P<tag>td|th)\b(?P<attrs>[^>]*)>", flags=re.IGNORECASE)
CATALOG_TABLE_BORDER_STYLE = "var(--catalog-table-border-style)"
CATALOG_TABLE_DATA_ATTR = "data-catalog-table"
CATALOG_TABLE_DATA_VALUE = "1"
CATALOG_TABLE_STYLE_MAP = {
    "border-style": CATALOG_TABLE_BORDER_STYLE,
    "width": "100%",
    "background-color": "white",
    "border-collapse": "collapse",
    "font-size": "14px",
}
CATALOG_TABLE_CELL_STYLE_MAP = {
    "padding": "12px",
    "border": "1px solid #e1e1e1",
}
REGULAR_TABLE_BORDER_STYLES = {
    "solid",
    "dotted",
    "dashed",
    "double",
    "groove",
    "ridge",
    "inset",
    "outset",
    "none",
    "hidden",
}
CATALOG_ONLY_TABLE_STYLE_MAP = {
    "background-color": "white",
    "font-size": "14px",
}
REMOTE_IMPORT_MAX_BYTES = 100 * 1024 * 1024
REMOTE_IMPORT_TIMEOUT_SECONDS = 30


def parse_html_attrs(attrs_text: str) -> dict[str, str]:
    return {match.group(1).lower(): match.group(2) for match in ATTR_RE.finditer(attrs_text or "")}


def parse_style_declarations(style_text: str | None) -> OrderedDict[str, str]:
    declarations = OrderedDict()
    for chunk in str(style_text or "").split(";"):
        if ":" not in chunk:
            continue
        key, value = chunk.split(":", 1)
        key = key.strip().lower()
        value = value.strip()
        if key and value:
            declarations[key] = value
    return declarations


def serialize_style_declarations(declarations: OrderedDict[str, str]) -> str:
    return "; ".join(f"{key}: {value}" for key, value in declarations.items())


def merge_style_declarations(existing: str | None, enforced: dict[str, str]) -> str:
    declarations = parse_style_declarations(existing)
    for key, value in enforced.items():
        declarations[key] = value
    return serialize_style_declarations(declarations)


def strip_style_declarations(existing: str | None, removable: dict[str, str]) -> str:
    declarations = parse_style_declarations(existing)
    for key, value in removable.items():
        if declarations.get(key) == value:
            declarations.pop(key, None)
    return serialize_style_declarations(declarations)


def serialize_html_attrs(attrs: dict[str, str]) -> str:
    return "".join(f' {key}="{value}"' for key, value in attrs.items() if value not in (None, ""))


def sanitize_catalog_tables(html: str | None) -> str | None:
    if not html:
        return html

    def replace_table(match):
        attrs = parse_html_attrs(match.group("attrs"))
        classes = [item for item in attrs.get("class", "").split() if item.strip()]
        table_style = parse_style_declarations(attrs.get("style"))
        border_style = table_style.get("border-style")
        has_catalog_class = "catalog-table" in classes
        has_catalog_border_style = border_style == CATALOG_TABLE_BORDER_STYLE
        has_regular_border_style = border_style in REGULAR_TABLE_BORDER_STYLES
        has_catalog_flag = attrs.get(CATALOG_TABLE_DATA_ATTR) == CATALOG_TABLE_DATA_VALUE
        has_catalog_origin = any((has_catalog_class, has_catalog_border_style, has_catalog_flag))
        explicit_non_catalog_style = has_catalog_origin and not any((has_catalog_border_style, has_catalog_class))

        if not any((has_catalog_origin, has_regular_border_style)):
            return match.group(0)

        classes = [item for item in classes if item != "catalog-table"]
        if classes:
            attrs["class"] = " ".join(classes).strip()
        else:
            attrs.pop("class", None)
        if has_catalog_border_style or has_catalog_class:
            attrs[CATALOG_TABLE_DATA_ATTR] = CATALOG_TABLE_DATA_VALUE
            attrs["style"] = merge_style_declarations(attrs.get("style"), CATALOG_TABLE_STYLE_MAP)
            attrs.pop("border", None)
            attrs.pop("cellpadding", None)
            attrs.pop("cellspacing", None)
        else:
            attrs.pop(CATALOG_TABLE_DATA_ATTR, None)
            stripped_table_style = strip_style_declarations(attrs.get("style"), CATALOG_ONLY_TABLE_STYLE_MAP)
            if has_regular_border_style:
                regular_table_style = parse_style_declarations(stripped_table_style)
                border_width = regular_table_style.get("border-width") or "1px"
                regular_table_style["border-style"] = border_style
                regular_table_style["border-width"] = border_width
                regular_table_style.setdefault("border-collapse", "collapse")
                attrs["style"] = serialize_style_declarations(regular_table_style)
                attrs.pop("border", None)
            elif stripped_table_style:
                attrs["style"] = stripped_table_style
            else:
                attrs.pop("style", None)

        def replace_row(row_match):
            row_attrs = parse_html_attrs(row_match.group("attrs"))
            row_content = row_match.group("content")
            cell_index = 0

            def replace_cell(cell_match):
                nonlocal cell_index
                cell_index += 1
                tag = cell_match.group("tag").lower()
                cell_attrs = parse_html_attrs(cell_match.group("attrs"))
                alignment = "left" if cell_index == 1 else "center"
                if has_catalog_border_style or has_catalog_class:
                    cell_attrs["style"] = merge_style_declarations(
                        cell_attrs.get("style"),
                        {
                            **CATALOG_TABLE_CELL_STYLE_MAP,
                            "text-align": alignment,
                        },
                    )
                elif has_regular_border_style:
                    stripped_cell_style = strip_style_declarations(
                        cell_attrs.get("style"),
                        {
                            **CATALOG_TABLE_CELL_STYLE_MAP,
                            "text-align": alignment,
                        },
                    )
                    if stripped_cell_style:
                        cell_attrs["style"] = stripped_cell_style
                    else:
                        cell_attrs.pop("style", None)
                    regular_table_style = parse_style_declarations(attrs.get("style"))
                    cell_attrs["style"] = merge_style_declarations(
                        cell_attrs.get("style"),
                        {
                            "border": (
                                f"{regular_table_style.get('border-width', '1px')} "
                                f"{regular_table_style.get('border-style', 'solid')} "
                                f"{regular_table_style.get('border-color', 'currentColor')}"
                            ),
                        },
                    )
                elif explicit_non_catalog_style:
                    stripped_cell_style = strip_style_declarations(
                        cell_attrs.get("style"),
                        {
                            **CATALOG_TABLE_CELL_STYLE_MAP,
                            "text-align": alignment,
                        },
                    )
                    if stripped_cell_style:
                        cell_attrs["style"] = stripped_cell_style
                    else:
                        cell_attrs.pop("style", None)
                else:
                    cell_attrs["style"] = merge_style_declarations(
                        cell_attrs.get("style"),
                        {
                            **CATALOG_TABLE_CELL_STYLE_MAP,
                            "text-align": alignment,
                        },
                    )
                return f"<{tag}{serialize_html_attrs(cell_attrs)}>"

            rebuilt_row_content = CELL_OPEN_RE.sub(replace_cell, row_content)
            return f"<tr{serialize_html_attrs(row_attrs)}>{rebuilt_row_content}</tr>"

        rebuilt_content = TR_BLOCK_RE.sub(replace_row, match.group("content"))
        return f"<table{serialize_html_attrs(attrs)}>{rebuilt_content}</table>"

    return TABLE_BLOCK_RE.sub(replace_table, html)


def ordered_product_media_queryset(product):
    return ProductMedia.objects.filter(product=product).order_by("-is_primary", "sort_order", "id")


def sync_product_media(product):
    ordered_urls = [media.url for media in ordered_product_media_queryset(product) if media.url]
    product.media = ordered_urls or None
    product.save(update_fields=["media"])


def ensure_single_primary(product):
    primary_media = list(ProductMedia.objects.filter(product=product, is_primary=True).order_by("sort_order", "id"))
    if not primary_media:
        first_media = ProductMedia.objects.filter(product=product).order_by("sort_order", "id").first()
        if first_media:
            first_media.is_primary = True
            first_media.save(update_fields=["is_primary"])
        return
    keeper = primary_media[0]
    ProductMedia.objects.filter(product=product, is_primary=True).exclude(pk=keeper.pk).update(is_primary=False)


def resolve_group(raw_value: str):
    if not raw_value:
        return None, False
    value = str(raw_value).strip()
    slug = transliterate_slug(value)
    group = Group.objects.filter(slug=slug).first() or Group.objects.filter(name=value).first()
    if group:
        return group, False
    return Group.objects.create(name=value, slug=slug), True


def resolve_brand(raw_value: str):
    if not raw_value:
        return None, False
    value = str(raw_value).strip()
    slug = transliterate_slug(value)
    brand = Brand.objects.filter(slug=slug).first() or Brand.objects.filter(name=value).first()
    if brand:
        return brand, False
    return Brand.objects.create(name=value, slug=slug), True


def resolve_characteristic(group: Group, header: str, sample_value):
    name = characteristic_name_from_header(header)
    slug = transliterate_slug(name)
    characteristic = Characteristic.objects.filter(group=group, slug=slug).first()
    if characteristic:
        updated = False
        inferred_type = infer_data_type(sample_value)
        if not characteristic.name:
            characteristic.name = name
            updated = True
        if not characteristic.data_type:
            characteristic.data_type = inferred_type
            updated = True
        if updated:
            characteristic.save(update_fields=["name", "data_type"])
        return characteristic, False
    return Characteristic.objects.create(
        group=group,
        name=name,
        slug=slug,
        data_type=infer_data_type(sample_value),
        is_filterable=True,
        is_searchable=False,
    ), True


PRODUCT_OPTIONAL_IMPORT_FIELDS = (
    "description",
    "assortment_html",
    "characteristics_html",
    "search_tsv",
    "seo_title",
    "seo_h1",
    "seo_description",
    "seo_keywords",
    "seo_canonical_url",
    "seo_robots",
)


def clean_optional_text(value):
    text = str(value or "").strip()
    return text or None


def sync_imported_product_media(product, values, warnings, row_number):
    ProductMedia.objects.filter(product=product).delete()
    created_items = []
    for index, value in enumerate(values):
        uploaded, warning = resolve_import_file_reference(value, "product_media")
        if warning:
            warnings.append(f"Row {row_number}: media skipped for SKU '{product.sku}'. {warning}")
            continue
        item = ProductMedia.objects.create(
            product=product,
            storage_path=uploaded["storage_path"],
            url=uploaded["url"],
            mime_type=uploaded["mime_type"],
            media_kind=infer_file_kind(uploaded["mime_type"]),
            size_bytes=uploaded["size_bytes"],
            is_primary=index == 0,
            sort_order=index,
            alt_text=product.name,
        )
        created_items.append(item)
    ensure_single_primary(product)
    sync_product_media(product)
    return len(created_items)


def sync_imported_product_gallery(product, values, title_values, warnings, row_number):
    ProductGalleryItem.objects.filter(product=product).delete()
    created_count = 0
    for index, value in enumerate(values):
        uploaded, warning = resolve_import_file_reference(value, "product_gallery")
        if warning:
            warnings.append(f"Row {row_number}: gallery file skipped for SKU '{product.sku}'. {warning}")
            continue
        ProductGalleryItem.objects.create(
            product=product,
            title=title_values[index] if index < len(title_values) else uploaded["title"],
            storage_path=uploaded["storage_path"],
            url=uploaded["url"],
            mime_type=uploaded["mime_type"],
            file_kind=infer_file_kind(uploaded["mime_type"]),
            size_bytes=uploaded["size_bytes"],
            sort_order=index,
        )
        created_count += 1
    return created_count


def sync_imported_titled_files(model, product, values, title_values, folder_name, warnings, row_number, warning_label):
    model.objects.filter(product=product).delete()
    created_count = 0
    for index, value in enumerate(values):
        uploaded, warning = resolve_import_file_reference(value, folder_name)
        if warning:
            warnings.append(f"Row {row_number}: {warning_label} skipped for SKU '{product.sku}'. {warning}")
            continue
        model.objects.create(
            product=product,
            title=title_values[index] if index < len(title_values) else uploaded["title"],
            storage_path=uploaded["storage_path"],
            url=uploaded["url"],
            mime_type=uploaded["mime_type"],
            size_bytes=uploaded["size_bytes"],
            sort_order=index,
        )
        created_count += 1
    return created_count


def import_products_from_workbook(workbook_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("openpyxl is not installed. Add it to the environment before importing XLSX.") from exc

    workbook = load_workbook(workbook_file, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration as exc:
        raise ValidationError("The XLSX file is empty.") from exc

    required_headers = {"sku", "name", "price", "currency"}
    missing = sorted(required_headers - set(headers))
    if missing:
        raise ValidationError(f"Missing required columns: {', '.join(missing)}")

    char_headers = [header for header in headers if header.startswith("char_")]
    counters = {
        "products_created": 0,
        "products_updated": 0,
        "groups_created": 0,
        "brands_created": 0,
        "characteristics_created": 0,
        "product_characteristics_upserted": 0,
        "media_items_imported": 0,
        "gallery_items_imported": 0,
        "certificates_imported": 0,
        "rows_skipped": 0,
    }
    warnings = []

    with transaction.atomic():
        for row_number, row in enumerate(rows, start=2):
            payload = dict(zip(headers, row))
            sku = str(payload.get("sku") or "").strip()
            if not sku:
                counters["rows_skipped"] += 1
                continue

            name = str(payload.get("name") or "").strip()
            if not name:
                warnings.append(f"Row {row_number}: skipped because product name is empty for SKU '{sku}'.")
                counters["rows_skipped"] += 1
                continue

            group_name = str(payload.get("group_slug") or "").strip()
            brand_name = str(payload.get("brand_slug") or "").strip()

            group, group_created = resolve_group(group_name) if group_name else (None, False)
            brand, brand_created = resolve_brand(brand_name) if brand_name else (None, False)
            counters["groups_created"] += int(group_created)
            counters["brands_created"] += int(brand_created)

            defaults = {
                "name": name,
                "price": parse_decimal(payload.get("price")),
                "currency": str(payload.get("currency") or "RUB").strip() or "RUB",
                "group": group,
                "brand": brand,
                "available": parse_bool(payload.get("available")),
            }
            for field_name in PRODUCT_OPTIONAL_IMPORT_FIELDS:
                if field_name in payload:
                    defaults[field_name] = clean_optional_text(payload.get(field_name))
            if "slug" in payload:
                defaults["slug"] = clean_optional_text(payload.get("slug")) or ""

            if "media_urls" in payload:
                defaults["media"] = split_media_urls(payload.get("media_urls"))

            product, created = Product.objects.update_or_create(sku=sku, defaults=defaults)
            counters["products_created"] += int(created)
            counters["products_updated"] += int(not created)

            if "media_urls" in payload:
                media_values = split_media_urls(payload.get("media_urls")) or []
                counters["media_items_imported"] += sync_imported_product_media(product, media_values, warnings, row_number)

            if "gallery_urls" in payload:
                counters["gallery_items_imported"] += sync_imported_product_gallery(
                    product,
                    split_media_urls(payload.get("gallery_urls")) or [],
                    split_title_values(payload.get("gallery_titles")),
                    warnings,
                    row_number,
                )

            if "certificate_urls" in payload:
                counters["certificates_imported"] += sync_imported_titled_files(
                    ProductCertificate,
                    product,
                    split_media_urls(payload.get("certificate_urls")) or [],
                    split_title_values(payload.get("certificate_titles")),
                    "product_certificates",
                    warnings,
                    row_number,
                    "certificate",
                )

            if not group and char_headers:
                has_characteristics = any(payload.get(header) not in (None, "") for header in char_headers)
                if has_characteristics:
                    warnings.append(f"Row {row_number}: characteristics for SKU '{sku}' were skipped because group_slug is empty.")
                continue

            for header in char_headers:
                raw_value = payload.get(header)
                if raw_value in (None, ""):
                    continue
                characteristic, char_created = resolve_characteristic(group, header, raw_value)
                counters["characteristics_created"] += int(char_created)
                ProductCharacteristic.objects.update_or_create(
                    product=product,
                    characteristic=characteristic,
                    defaults={"value": str(raw_value).strip()},
                )
                counters["product_characteristics_upserted"] += 1

    return counters, warnings


def characteristic_header_from_name(name: str) -> str:
    normalized = str(name or "").strip().replace("\n", " ")
    normalized = re.sub(r"\s+", "_", normalized)
    return f"char_{normalized}"


def build_product_export_rows(products_queryset):
    products = list(
        products_queryset.select_related("group", "brand").prefetch_related(
            "characteristics__characteristic",
            "media_files",
            "gallery_items",
            "certificates",
        )
    )
    characteristic_names = {}
    for product in products:
        for product_characteristic in product.characteristics.all():
            characteristic_names[product_characteristic.characteristic.slug] = product_characteristic.characteristic.name

    ordered_characteristics = sorted(characteristic_names.items(), key=lambda item: item[1].lower())
    char_headers = [characteristic_header_from_name(name) for _, name in ordered_characteristics]

    headers = [
        "sku",
        "slug",
        "name",
        "price",
        "currency",
        "description",
        "assortment_html",
        "characteristics_html",
        "search_tsv",
        "seo_title",
        "seo_h1",
        "seo_description",
        "seo_keywords",
        "seo_canonical_url",
        "seo_robots",
        "group_slug",
        "brand_slug",
        "available",
        "media_urls",
        "gallery_urls",
        "gallery_titles",
        "certificate_urls",
        "certificate_titles",
        *char_headers,
    ]
    rows = []
    for product in products:
        values_by_slug = {
            product_characteristic.characteristic.slug: product_characteristic.value
            for product_characteristic in product.characteristics.all()
        }
        product_media_urls = [item.url for item in product.media_files.all()]
        rows.append([
            product.sku,
            product.slug,
            product.name,
            str(product.price),
            product.currency,
            product.description or "",
            product.assortment_html or "",
            product.characteristics_html or "",
            product.search_tsv or "",
            product.seo_title or "",
            product.seo_h1 or "",
            product.seo_description or "",
            product.seo_keywords or "",
            product.seo_canonical_url or "",
            product.seo_robots or "",
            product.group.name if product.group else "",
            product.brand.name if product.brand else "",
            "Да" if product.available else "Нет",
            ",".join(product_media_urls or (product.media or [])),
            ",".join(item.url for item in product.gallery_items.all()),
            ",".join(item.title or "" for item in product.gallery_items.all()),
            ",".join(item.url for item in product.certificates.all()),
            ",".join(item.title or "" for item in product.certificates.all()),
            *[
                values_by_slug.get(characteristic_slug, "") or ""
                for characteristic_slug, _ in ordered_characteristics
            ],
        ])
    return headers, rows


def workbook_bytes_from_headers_and_rows(headers, rows, title="Products"):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ValidationError("openpyxl is not installed. Add it to the environment before exporting XLSX.") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title[:31] or "Products"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


SEO_FIELD_NAMES = (
    "seo_title",
    "seo_h1",
    "seo_description",
    "seo_keywords",
    "seo_canonical_url",
    "seo_robots",
)
SEO_AUTO_HELP_TEXT = "Leave empty to let the API generate this SEO value automatically."
SEO_GROUP_PLACEHOLDER_HELP_TEXT = (
    "Supports placeholders: {name}, {slug}, {parent}, {city}, {city_slug}, {city_prep}."
)


class AdminMediaFormMixin(forms.ModelForm):
    media_upload = forms.FileField(required=False, label="Upload file")
    media_field_name = None
    upload_folder_name = "generic"

    def save(self, commit=True):
        instance = super().save(commit=False)
        upload = self.cleaned_data.get("media_upload")
        if upload:
            uploaded = save_admin_upload(upload, self.upload_folder_name)
            setattr(instance, self.media_field_name, uploaded["url"])
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class HtmlTableSanitizerMixin:
    html_field_names = ()

    def clean(self):
        cleaned_data = super().clean()
        for field_name in self.html_field_names:
            if field_name in cleaned_data:
                cleaned_data[field_name] = sanitize_catalog_tables(cleaned_data.get(field_name))
        return cleaned_data


def render_multiline_text(value):
    text = escape(str(value or ""))
    return text.replace("\n", "<br>")


class SeoFieldsAdminFormMixin:
    seo_optional_fields = SEO_FIELD_NAMES
    seo_help_text_by_field = {}
    seo_default_help_text = None

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        for field_name in self.seo_optional_fields:
            field = self.fields.get(field_name)
            if not field:
                continue
            field.required = False
            help_text = self.seo_help_text_by_field.get(field_name, self.seo_default_help_text)
            if help_text:
                field.help_text = help_text


class UploadedAssetAdminFormMixin(forms.ModelForm):
    upload_field_name = None
    upload_folder_name = "generic"
    generated_optional_fields = ()
    inferred_kind_field_name = None
    title_field_name = None

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        mark_generated_file_fields_optional(self, extra_fields=list(self.generated_optional_fields))

    def clean(self):
        cleaned_data = super().clean()
        validate_new_file_upload(self, self.upload_field_name)
        return cleaned_data

    def save(self, commit=True):
        instance = super().save(commit=False)
        upload = self.cleaned_data.get(self.upload_field_name)
        if upload:
            uploaded = save_admin_upload(upload, self.upload_folder_name)
            instance.storage_path = uploaded["storage_path"]
            instance.url = uploaded["url"]
            instance.mime_type = uploaded["mime_type"]
            instance.size_bytes = uploaded["size_bytes"]
            if self.inferred_kind_field_name:
                setattr(instance, self.inferred_kind_field_name, infer_file_kind(uploaded["mime_type"]))
            if self.title_field_name and not getattr(instance, self.title_field_name):
                setattr(instance, self.title_field_name, upload.name)
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class TabbedFieldsetsAdminMixin:
    class Media:
        css = {"all": ("shop/css/admin_tabbed_fieldsets.css",)}
        js = ("shop/js/admin_tabbed_fieldsets.js",)


class PublishWorkflowAdminMixin:
    preview_template_name = "admin/shop/content_preview.html"
    singleton_publication = False
    actions = ("publish_selected", "move_to_draft")

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "<path:object_id>/preview/",
                self.admin_site.admin_view(self.preview_view),
                name=f"{self.model._meta.app_label}_{self.model._meta.model_name}_preview",
            ),
        ]
        return custom_urls + urls

    def save_model(self, request, obj, form, change):
        previous_status = None
        if change:
            previous = self.model.objects.filter(pk=obj.pk).only("status").first()
            previous_status = previous.status if previous else None

        if hasattr(obj, "updated_by") and request.user.is_authenticated:
            obj.updated_by = request.user

        if hasattr(obj, "published_at") and obj.status == PUBLISH_STATUS_PUBLISHED:
            if previous_status != PUBLISH_STATUS_PUBLISHED or not obj.published_at:
                obj.published_at = timezone.now()

        super().save_model(request, obj, form, change)
        if obj.status == PUBLISH_STATUS_PUBLISHED:
            self._unpublish_siblings(obj, request.user if request.user.is_authenticated else None)

    def _unpublish_siblings(self, obj, user):
        if not self.singleton_publication:
            return
        siblings = self.model.objects.exclude(pk=obj.pk).filter(status=PUBLISH_STATUS_PUBLISHED)
        for sibling in siblings:
            sibling.status = PUBLISH_STATUS_DRAFT
            if hasattr(sibling, "updated_by"):
                sibling.updated_by = user
            sibling.save()

    @admin.action(description="Publish selected items")
    def publish_selected(self, request, queryset):
        if self.singleton_publication and queryset.count() > 1:
            self.message_user(request, "Select only one record to publish for this section.", level=messages.ERROR)
            return

        published_count = 0
        for obj in queryset:
            obj.status = PUBLISH_STATUS_PUBLISHED
            if hasattr(obj, "updated_by") and request.user.is_authenticated:
                obj.updated_by = request.user
            if hasattr(obj, "published_at"):
                obj.published_at = timezone.now()
            obj.save()
            self._unpublish_siblings(obj, request.user if request.user.is_authenticated else None)
            published_count += 1

        self.message_user(request, f"Published items: {published_count}.", level=messages.SUCCESS)

    @admin.action(description="Move selected items to draft")
    def move_to_draft(self, request, queryset):
        updated_count = 0
        for obj in queryset:
            obj.status = PUBLISH_STATUS_DRAFT
            if hasattr(obj, "updated_by") and request.user.is_authenticated:
                obj.updated_by = request.user
            obj.save()
            updated_count += 1

        self.message_user(request, f"Draft items: {updated_count}.", level=messages.SUCCESS)

    @admin.display(description="Preview")
    def preview_link(self, obj):
        if not obj or not obj.pk:
            return "Save to preview"
        url = reverse(f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_preview", args=[obj.pk])
        return format_html('<a href="{}" target="_blank">open preview</a>', url)

    def preview_view(self, request, object_id):
        obj = self.get_object(request, object_id)
        if obj is None:
            raise PermissionDenied("Object not found.")
        if not self.has_view_or_change_permission(request, obj):
            raise PermissionDenied("You do not have permission to preview this object.")

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "object": obj,
            "title": f"Preview: {obj}",
            "preview_html": mark_safe(self.render_preview_html(obj)),
            "back_url": reverse(f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_change", args=[obj.pk]),
            "history_url": reverse(f"admin:{self.model._meta.app_label}_{self.model._meta.model_name}_history", args=[obj.pk]),
        }
        return TemplateResponse(request, self.preview_template_name, context)

    def render_preview_html(self, obj):
        return f"<pre>{escape(str(obj))}</pre>"


class ProductImportForm(forms.Form):
    xlsx_file = forms.FileField(label="XLSX file")

    def clean_xlsx_file(self):
        file = self.cleaned_data["xlsx_file"]
        if not file.name.lower().endswith(".xlsx"):
            raise ValidationError("Upload an .xlsx file.")
        return file


class ProductExportForm(forms.Form):
    group = forms.ModelChoiceField(
        queryset=Group.objects.order_by("name"),
        required=False,
        label="Group",
        help_text="Choose a group to export a single XLSX. Leave empty to download a ZIP with separate files for each group.",
    )


class SynonymListField(forms.JSONField):
    widget = forms.Textarea

    def __init__(self, *args, **kwargs):
        kwargs.setdefault("required", False)
        kwargs.setdefault("help_text", "One synonym per line. The value is stored as a JSON list.")
        super().__init__(*args, **kwargs)

    def prepare_value(self, value):
        if isinstance(value, list):
            return "\n".join(str(item) for item in value if str(item).strip())
        return super().prepare_value(value)

    def to_python(self, value):
        if value in self.empty_values:
            return []
        if isinstance(value, str) and not value.lstrip().startswith("["):
            return [line.strip() for line in value.splitlines() if line.strip()]
        parsed = super().to_python(value)
        if parsed in self.empty_values:
            return []
        if not isinstance(parsed, list):
            raise ValidationError("Enter a list of synonyms.")
        return [str(item).strip() for item in parsed if str(item).strip()]


class BrandAdminForm(AdminMediaFormMixin):
    media_field_name = "media"
    upload_folder_name = "brands"
    search_synonyms = SynonymListField(label="Search synonyms", required=False)

    class Meta:
        model = Brand
        fields = "__all__"


class GroupAdminForm(SeoFieldsAdminFormMixin, AdminMediaFormMixin):
    media_field_name = "media"
    upload_folder_name = "groups"
    search_synonyms = SynonymListField(label="Search synonyms", required=False)
    seo_help_text_by_field = {
        "seo_title": SEO_GROUP_PLACEHOLDER_HELP_TEXT,
        "seo_h1": SEO_GROUP_PLACEHOLDER_HELP_TEXT,
        "seo_description": SEO_GROUP_PLACEHOLDER_HELP_TEXT,
        "seo_keywords": SEO_GROUP_PLACEHOLDER_HELP_TEXT,
        "seo_canonical_url": SEO_GROUP_PLACEHOLDER_HELP_TEXT,
    }

    class Meta:
        model = Group
        fields = "__all__"


class ProductAdminForm(HtmlTableSanitizerMixin, SeoFieldsAdminFormMixin, AdminMediaFormMixin):
    media_field_name = "media"
    upload_folder_name = "products"
    html_field_names = ("assortment_html", "characteristics_html")
    seo_default_help_text = SEO_AUTO_HELP_TEXT

    class Meta:
        model = Product
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["search_tsv"].label = "Search synonyms"
        self.fields["search_tsv"].help_text = (
            "Comma-separated search synonyms and semantic hints. "
            "Example: огнезащита, огнезащитный материал, fireproof, огнезащ."
        )
        self.fields["assortment_html"].widget = TinyMCE(
            attrs={"cols": 120, "rows": 22},
            mce_attrs={"height": 460},
        )
        self.fields["characteristics_html"].widget = TinyMCE(
            attrs={"cols": 120, "rows": 22},
            mce_attrs={"height": 460},
        )


class NewsAdminForm(HtmlTableSanitizerMixin, AdminMediaFormMixin):
    media_field_name = "media"
    upload_folder_name = "news"
    html_field_names = ("content",)

    class Meta:
        model = News
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["content"].widget = TinyMCE(
            attrs={"cols": 120, "rows": 24},
            mce_attrs={"height": 520},
        )


class HtmlContentAdminForm(HtmlTableSanitizerMixin, forms.ModelForm):
    html_field_names = ("html_first", "html_second")

    class Meta:
        model = HtmlContent
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["html_first"].widget = TinyMCE(
            attrs={"cols": 120, "rows": 18},
            mce_attrs={"height": 420},
        )
        self.fields["html_second"].widget = TinyMCE(
            attrs={"cols": 120, "rows": 18},
            mce_attrs={"height": 420},
        )


class OrderEmailSettingsAdminForm(HtmlTableSanitizerMixin, forms.ModelForm):
    html_field_names = ("intro_html", "body_html", "footer_html")

    class Meta:
        model = OrderEmailSettings
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["intro_html"].widget = TinyMCE(
            attrs={"cols": 120, "rows": 16},
            mce_attrs={"height": 340},
        )
        self.fields["body_html"].widget = TinyMCE(
            attrs={"cols": 120, "rows": 18},
            mce_attrs={"height": 380},
        )
        self.fields["body_html"].help_text = (
            "For order templates: {{order_id}}, {{name}}, {{phone}}, {{email}}, {{address}}, "
            "{{comment}}, {{total_items}}, {{items_table}}, {{items_text}}. "
            "For inquiry templates: {{inquiry_id}}, {{name}}, {{phone}}, {{email}}, {{message}}, {{created_at}}."
        )
        self.fields["footer_html"].widget = TinyMCE(
            attrs={"cols": 120, "rows": 14},
            mce_attrs={"height": 280},
        )


class SliderAdminForm(forms.ModelForm):
    image_upload = forms.FileField(required=False, label="Upload image")

    class Meta:
        model = Slider
        fields = "__all__"

    def save(self, commit=True):
        instance = super().save(commit=False)
        upload = self.cleaned_data.get("image_upload")
        if upload:
            uploaded = save_admin_upload(upload, "slider")
            instance.image = uploaded["url"]
        if commit:
            instance.save()
            self.save_m2m()
        return instance


class ProductMediaAdminForm(UploadedAssetAdminFormMixin):
    media_upload = forms.FileField(required=False, label="Upload file")
    upload_field_name = "media_upload"
    upload_folder_name = "product_media"
    generated_optional_fields = ("media_kind",)
    inferred_kind_field_name = "media_kind"

    class Meta:
        model = ProductMedia
        fields = "__all__"


class ProductCertificateAdminForm(UploadedAssetAdminFormMixin):
    certificate_upload = forms.FileField(required=False, label="Upload certificate")
    upload_field_name = "certificate_upload"
    upload_folder_name = "product_certificates"
    title_field_name = "title"

    class Meta:
        model = ProductCertificate
        fields = "__all__"


class ProductGalleryItemAdminForm(UploadedAssetAdminFormMixin):
    gallery_upload = forms.FileField(required=False, label="Upload gallery file")
    upload_field_name = "gallery_upload"
    upload_folder_name = "product_gallery"
    generated_optional_fields = ("file_kind",)
    inferred_kind_field_name = "file_kind"
    title_field_name = "title"

    class Meta:
        model = ProductGalleryItem
        fields = "__all__"


class ProductCharacteristicAdminForm(forms.ModelForm):
    class Meta:
        model = ProductCharacteristic
        fields = "__all__"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        product = self.initial.get("product") or getattr(self.instance, "product", None)
        product_id = getattr(product, "pk", product)
        if not product_id and self.data:
            product_id = self.data.get("product")
        if product_id:
            selected_product = Product.objects.filter(pk=product_id).first()
            if selected_product and selected_product.group_id:
                self.fields["characteristic"].queryset = Characteristic.objects.filter(group=selected_product.group).order_by("name", "id")


class NewsAttachmentAdminForm(UploadedAssetAdminFormMixin):
    attachment_upload = forms.FileField(required=False, label="Upload attachment")
    upload_field_name = "attachment_upload"
    upload_folder_name = "news_attachments"
    title_field_name = "title"

    class Meta:
        model = NewsAttachment
        fields = "__all__"


class SertAdminForm(UploadedAssetAdminFormMixin):
    file_upload = forms.FileField(required=False, label="Upload file")
    upload_field_name = "file_upload"
    upload_folder_name = "serts"
    title_field_name = "title"

    class Meta:
        model = Sert
        fields = "__all__"


class ProductMediaInlineForm(ProductMediaAdminForm):
    class Meta(ProductMediaAdminForm.Meta):
        model = ProductMedia
        fields = "__all__"


class ProductMediaInline(admin.TabularInline):
    model = ProductMedia
    form = ProductMediaInlineForm
    extra = 1
    fields = (
        "media_upload",
        "preview",
        "media_kind",
        "mime_type",
        "size_bytes",
        "is_primary",
        "sort_order",
        "alt_text",
    )
    readonly_fields = ("preview", "media_kind", "mime_type", "size_bytes")
    ordering = ("-is_primary", "sort_order", "id")

    @admin.display(description="Preview")
    def preview(self, obj):
        if not obj or not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">open</a><br><img src="{0}" style="max-height:100px;max-width:140px;" />', obj.url)


class ProductCertificateInline(admin.TabularInline):
    model = ProductCertificate
    form = ProductCertificateAdminForm
    extra = 1
    fields = (
        "certificate_upload",
        "title",
        "document_link",
        "mime_type",
        "size_bytes",
        "sort_order",
    )
    readonly_fields = ("document_link", "mime_type", "size_bytes")
    ordering = ("sort_order", "id")

    @admin.display(description="Certificate")
    def document_link(self, obj):
        if not obj or not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">{1}</a>', obj.url, obj.title or "open")


class ProductGalleryItemInline(admin.TabularInline):
    model = ProductGalleryItem
    form = ProductGalleryItemAdminForm
    extra = 1
    fields = (
        "gallery_upload",
        "title",
        "file_kind",
        "document_link",
        "mime_type",
        "size_bytes",
        "sort_order",
    )
    readonly_fields = ("file_kind", "document_link", "mime_type", "size_bytes")
    ordering = ("sort_order", "id")

    @admin.display(description="Gallery file")
    def document_link(self, obj):
        if not obj or not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">{1}</a>', obj.url, obj.title or "open")


class NewsAttachmentInline(admin.TabularInline):
    model = NewsAttachment
    form = NewsAttachmentAdminForm
    extra = 1
    fields = (
        "attachment_upload",
        "title",
        "document_link",
        "mime_type",
        "size_bytes",
        "sort_order",
    )
    readonly_fields = ("document_link", "mime_type", "size_bytes")
    ordering = ("sort_order", "id")

    @admin.display(description="Attachment")
    def document_link(self, obj):
        if not obj or not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">{1}</a>', obj.url, obj.title or "open")


class PublicOrderItemInline(admin.TabularInline):
    model = PublicOrderItem
    extra = 1
    fields = ("product", "qty")
    autocomplete_fields = ("product",)
    can_delete = True


class MediaPreviewAdminMixin:
    @admin.display(description="Preview")
    def media_preview(self, obj):
        url = extract_media_url(getattr(obj, "media", None))
        if not url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">open</a><br><img src="{0}" style="max-height:120px;max-width:180px;" />', url)


@admin.register(Brand)
class BrandAdmin(MediaPreviewAdminMixin, admin.ModelAdmin):
    form = BrandAdminForm
    list_display = ("id", "name", "slug", "media_preview")
    search_fields = ("name", "slug")
    readonly_fields = ("media_preview",)
    fields = ("name", "slug", "search_synonyms", "media", "media_upload", "media_preview")


@admin.register(City)
class CityAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "slug", "name_in_prepositional", "sort_order", "is_active")
    search_fields = ("name", "slug", "name_in_prepositional")
    list_filter = ("is_active",)
    ordering = ("sort_order", "name", "id")
    fields = ("name", "slug", "name_in_prepositional", "sort_order", "is_active")


@admin.register(Group)
class GroupAdmin(TabbedFieldsetsAdminMixin, MediaPreviewAdminMixin, admin.ModelAdmin):
    form = GroupAdminForm
    list_display = ("id", "name", "slug", "parent", "media_preview")
    search_fields = ("name", "slug", "seo_title", "seo_h1")
    list_filter = ("parent",)
    readonly_fields = ("media_preview",)
    fieldsets = (
        ("Основное", {
            "classes": ("tabbed-fieldset",),
            "fields": (
                "parent",
                "name",
                "slug",
                "search_synonyms",
                "description",
                "media",
                "media_upload",
                "media_preview",
            ),
        }),
        ("SEO", {
            "classes": ("tabbed-fieldset",),
            "fields": SEO_FIELD_NAMES,
        }),
    )


@admin.register(Product)
class ProductAdmin(TabbedFieldsetsAdminMixin, MediaPreviewAdminMixin, admin.ModelAdmin):
    form = ProductAdminForm
    change_list_template = "admin/shop/product/change_list.html"
    inlines = [ProductMediaInline, ProductGalleryItemInline, ProductCertificateInline]
    list_display = ("id", "name", "sku", "brand", "group", "price", "available", "media_preview")
    search_fields = ("name", "sku", "slug", "search_tsv", "seo_title", "seo_h1")
    autocomplete_fields = ("brand", "group")
    list_filter = ("available", "brand", "group")
    readonly_fields = ("media_preview",)
    fieldsets = (
        ("Основное", {
            "classes": ("tabbed-fieldset",),
            "fields": (
                "sku",
                "slug",
                "name",
                "price",
                "currency",
                "description",
                "assortment_html",
                "characteristics_html",
                "group",
                "brand",
                "media_preview",
                "available",
                "search_tsv",
            ),
        }),
        ("SEO", {
            "classes": ("tabbed-fieldset",),
            "fields": SEO_FIELD_NAMES,
        }),
    )

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "import-xlsx/",
                self.admin_site.admin_view(self.import_xlsx_view),
                name="shop_product_import_xlsx",
            ),
            path(
                "export-xlsx/",
                self.admin_site.admin_view(self.export_xlsx_view),
                name="shop_product_export_xlsx",
            ),
        ]
        return custom_urls + urls

    def import_xlsx_view(self, request):
        if not self.has_add_permission(request):
            raise PermissionDenied("You do not have permission to import products.")

        form = ProductImportForm(request.POST or None, request.FILES or None)

        if request.method == "POST" and form.is_valid():
            try:
                counters, warnings = import_products_from_workbook(form.cleaned_data["xlsx_file"])
            except ValidationError as exc:
                form.add_error("xlsx_file", exc)
            else:
                self.message_user(
                    request,
                    (
                        "Import finished. "
                        f"Products created: {counters['products_created']}, "
                        f"updated: {counters['products_updated']}, "
                        f"groups created: {counters['groups_created']}, "
                        f"brands created: {counters['brands_created']}, "
                        f"characteristics created: {counters['characteristics_created']}, "
                        f"product characteristics upserted: {counters['product_characteristics_upserted']}, "
                        f"media imported: {counters['media_items_imported']}, "
                        f"gallery imported: {counters['gallery_items_imported']}, "
                        f"certificates imported: {counters['certificates_imported']}, "
                        f"rows skipped: {counters['rows_skipped']}."
                    ),
                    level=messages.SUCCESS,
                )
                for warning in warnings[:20]:
                    self.message_user(request, warning, level=messages.WARNING)
                if len(warnings) > 20:
                    self.message_user(request, f"Additional warnings not shown: {len(warnings) - 20}.", level=messages.WARNING)
                changelist_url = reverse("admin:shop_product_changelist")
                return HttpResponseRedirect(changelist_url)

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Import products from XLSX",
            "form": form,
            "subtitle": "Upload a spreadsheet with product fields, file/link columns, and char_* columns.",
        }
        return TemplateResponse(request, "admin/shop/product/import_xlsx.html", context)

    def export_xlsx_view(self, request):
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied("You do not have permission to export products.")

        form = ProductExportForm(request.POST or None)
        if request.method == "POST" and form.is_valid():
            selected_group = form.cleaned_data["group"]
            if selected_group:
                return self._export_single_group_workbook(selected_group)
            return self._export_grouped_zip()

        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Export products to XLSX",
            "form": form,
            "subtitle": "Download a spreadsheet in the same format as the import template. Pick one group for a single XLSX or leave it empty for a ZIP split by groups.",
            "submit_label": "Export",
        }
        return TemplateResponse(request, "admin/shop/product/export_xlsx.html", context)

    def _export_single_group_workbook(self, group):
        headers, rows = build_product_export_rows(Product.objects.filter(group=group).order_by("name", "id"))
        payload = workbook_bytes_from_headers_and_rows(headers, rows, title=group.slug or group.name)
        filename = f"products_{transliterate_slug(group.slug or group.name)}.xlsx"
        response = HttpResponse(
            payload,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    def _export_grouped_zip(self):
        buffer = BytesIO()
        with ZipFile(buffer, "w", compression=ZIP_DEFLATED) as archive:
            groups = Group.objects.order_by("name")
            for group in groups:
                headers, rows = build_product_export_rows(Product.objects.filter(group=group).order_by("name", "id"))
                if not rows:
                    continue
                archive.writestr(
                    f"products_{transliterate_slug(group.slug or group.name)}.xlsx",
                    workbook_bytes_from_headers_and_rows(headers, rows, title=group.slug or group.name),
                )

            headers, rows = build_product_export_rows(Product.objects.filter(group__isnull=True).order_by("name", "id"))
            if rows:
                archive.writestr(
                    "products_without_group.xlsx",
                    workbook_bytes_from_headers_and_rows(headers, rows, title="Without group"),
                )

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/zip")
        response["Content-Disposition"] = 'attachment; filename="products_by_groups.zip"'
        return response

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        ensure_single_primary(form.instance)
        sync_product_media(form.instance)


@admin.register(News)
class NewsAdmin(PublishWorkflowAdminMixin, MediaPreviewAdminMixin, admin.ModelAdmin):
    form = NewsAdminForm
    inlines = [NewsAttachmentInline]
    list_display = ("id", "title", "slug", "status", "published_at", "updated_at", "updated_by", "preview_link")
    search_fields = ("title", "slug")
    list_filter = ("status",)
    readonly_fields = ("media_preview", "created_at", "updated_at", "updated_by", "preview_link")
    fields = (
        "title",
        "slug",
        "content",
        "media",
        "media_upload",
        "media_preview",
        "status",
        "published_at",
        "preview_link",
        "updated_by",
        "created_at",
        "updated_at",
    )

    def render_preview_html(self, obj):
        media_url = extract_media_url(obj.media)
        media_block = ""
        if media_url:
            media_block = f'<p><img src="{escape(media_url)}" style="max-width: 100%; height: auto;" /></p>'
        published_line = f"<p><strong>Published at:</strong> {escape(obj.published_at or 'Not published')}</p>"
        return (
            f"<article>{media_block}<h1>{escape(obj.title)}</h1>{published_line}{obj.content}</article>"
        )


@admin.register(Slider)
class SliderAdmin(PublishWorkflowAdminMixin, admin.ModelAdmin):
    form = SliderAdminForm
    list_display = ("id", "title", "slug", "sort_order", "status", "updated_at", "updated_by", "preview_link")
    list_filter = ("status",)
    search_fields = ("title", "slug", "text")
    readonly_fields = ("image", "image_preview", "preview_link", "created_at", "updated_at", "updated_by")
    fields = (
        "title",
        "slug",
        "text",
        "image_upload",
        "image",
        "image_preview",
        "sort_order",
        "status",
        "preview_link",
        "updated_by",
        "created_at",
        "updated_at",
    )

    @admin.display(description="Preview")
    def image_preview(self, obj):
        if not obj or not obj.image:
            return "No file"
        return format_html('<a href="{0}" target="_blank">open</a><br><img src="{0}" style="max-height:120px;max-width:180px;" />', obj.image)

    def render_preview_html(self, obj):
        image_block = ""
        if obj.image:
            image_block = f'<p><img src="{escape(obj.image)}" style="max-width: 100%; height: auto;" /></p>'
        text_block = f"<div>{render_multiline_text(obj.text)}</div>" if obj.text else ""
        return f"<section>{image_block}<h1>{escape(obj.title)}</h1>{text_block}</section>"


@admin.register(Inquiry)
class InquiryAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "phone", "email", "created_at")
    search_fields = ("name", "phone", "email", "message")
    readonly_fields = ("created_at",)
    fields = ("name", "phone", "email", "message", "created_at")


@admin.register(HtmlContent)
class HtmlContentAdmin(PublishWorkflowAdminMixin, admin.ModelAdmin):
    form = HtmlContentAdminForm
    singleton_publication = True
    list_display = ("id", "title", "status", "updated_at", "updated_by", "preview_link")
    list_filter = ("status",)
    search_fields = ("title", "html_first", "html_second")
    readonly_fields = ("preview_link", "created_at", "updated_at", "updated_by")
    fields = (
        "title",
        "html_first",
        "html_second",
        "status",
        "preview_link",
        "updated_by",
        "created_at",
        "updated_at",
    )

    def render_preview_html(self, obj):
        return f"<section>{obj.html_first}<hr>{obj.html_second}</section>"


@admin.register(ContactInfo)
class ContactInfoAdmin(PublishWorkflowAdminMixin, admin.ModelAdmin):
    singleton_publication = True
    list_display = ("id", "title", "phone", "email", "status", "updated_at", "updated_by", "preview_link")
    list_filter = ("status",)
    search_fields = ("title", "full_name", "address", "schedule", "phone", "email", "yandex_link", "gis_link")
    readonly_fields = ("preview_link", "created_at", "updated_at", "updated_by")
    fields = (
        "title",
        "full_name",
        "address",
        "latitude",
        "longitude",
        "yandex_link",
        "gis_link",
        "schedule",
        "phone",
        "email",
        "status",
        "preview_link",
        "updated_by",
        "created_at",
        "updated_at",
    )

    def render_preview_html(self, obj):
        full_name_block = f"<p><strong>Contact person:</strong> {escape(obj.full_name)}</p>" if obj.full_name else ""
        coordinates_block = (
            f"<p><strong>Coordinates:</strong> {escape(obj.latitude)} / {escape(obj.longitude)}</p>"
            if obj.latitude is not None and obj.longitude is not None
            else ""
        )
        return (
            f"<section><h1>{escape(obj.title)}</h1>"
            f"{full_name_block}"
            f"<p><strong>Address:</strong><br>{render_multiline_text(obj.address)}</p>"
            f"{coordinates_block}"
            f"<p><strong>Yandex:</strong> {escape(obj.yandex_link or '')}</p>"
            f"<p><strong>2GIS:</strong> {escape(obj.gis_link or '')}</p>"
            f"<p><strong>Schedule:</strong><br>{render_multiline_text(obj.schedule)}</p>"
            f"<p><strong>Phone:</strong> {escape(obj.phone)}</p>"
            f"<p><strong>Email:</strong> {escape(obj.email)}</p></section>"
        )


@admin.register(Agent)
class AgentAdmin(PublishWorkflowAdminMixin, admin.ModelAdmin):
    list_display = ("id", "full_name", "position", "email", "phone", "sort_order", "status", "updated_at", "updated_by", "preview_link")
    list_filter = ("status",)
    search_fields = ("full_name", "position", "email", "phone")
    readonly_fields = ("preview_link", "created_at", "updated_at", "updated_by")
    fields = (
        "full_name",
        "position",
        "email",
        "phone",
        "sort_order",
        "status",
        "preview_link",
        "updated_by",
        "created_at",
        "updated_at",
    )

    def render_preview_html(self, obj):
        return (
            f"<section><h1>{escape(obj.full_name)}</h1>"
            f"<p>{escape(obj.position)}</p>"
            f"<p><strong>Phone:</strong> {escape(obj.phone)}</p>"
            f"<p><strong>Email:</strong> {escape(obj.email)}</p></section>"
        )


@admin.register(PublicOrder)
class PublicOrderAdmin(admin.ModelAdmin):
    inlines = [PublicOrderItemInline]
    list_display = ("id", "name", "phone", "email", "status", "total_items", "created_at")
    list_filter = ("status", "created_at")
    search_fields = ("name", "phone", "email", "address", "comment", "items__product__name", "items__product__sku")
    readonly_fields = ("created_at", "total_items")
    fields = ("name", "phone", "email", "address", "comment", "status", "total_items", "created_at")
    autocomplete_fields = ()

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        order = form.instance
        total_items = sum(order.items.values_list("qty", flat=True))
        if order.total_items != total_items:
            order.total_items = total_items
            order.save(update_fields=["total_items"])


@admin.register(OrderEmailRecipient)
class OrderEmailRecipientAdmin(admin.ModelAdmin):
    list_display = ("id", "email", "name", "is_active", "updated_at")
    list_filter = ("is_active",)
    search_fields = ("email", "name")
    readonly_fields = ("created_at", "updated_at")
    fields = ("email", "name", "is_active", "created_at", "updated_at")


@admin.register(OrderEmailSettings)
class OrderEmailSettingsAdmin(PublishWorkflowAdminMixin, admin.ModelAdmin):
    form = OrderEmailSettingsAdminForm
    singleton_publication = True
    list_display = ("id", "title", "notification_type", "subject", "status", "updated_at", "updated_by", "preview_link")
    list_filter = ("status",)
    search_fields = ("title", "subject", "notification_type", "intro_html", "footer_html")
    readonly_fields = ("preview_link", "created_at", "updated_at", "updated_by")
    fields = (
        "title",
        "notification_type",
        "subject",
        "intro_html",
        "body_html",
        "footer_html",
        "status",
        "preview_link",
        "updated_by",
        "created_at",
        "updated_at",
    )

    def _unpublish_siblings(self, obj, user):
        siblings = (
            self.model.objects.exclude(pk=obj.pk)
            .filter(status=PUBLISH_STATUS_PUBLISHED, notification_type=obj.notification_type)
        )
        for sibling in siblings:
            sibling.status = PUBLISH_STATUS_DRAFT
            if hasattr(sibling, "updated_by"):
                sibling.updated_by = user
            sibling.save()

    def render_preview_html(self, obj):
        return build_notification_preview_html(obj.notification_type, obj)


@admin.register(MediaLibrary)
class MediaLibraryAdmin(admin.ModelAdmin):
    change_list_template = "admin/shop/medialibrary/change_list.html"
    actions = None

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "delete-asset/",
                self.admin_site.admin_view(self.delete_asset_view),
                name="shop_medialibrary_delete_asset",
            ),
        ]
        return custom_urls + urls

    def changelist_view(self, request, extra_context=None):
        if not self.has_view_or_change_permission(request):
            raise PermissionDenied("You do not have permission to view the media library.")
        search_query = request.GET.get("q", "").strip()
        usage_filter = request.GET.get("usage", "").strip()
        usage_options = [
            "Brand image",
            "Group image",
            "Slider image",
            "Product media field",
            "News media field",
            "Product media",
            "Product gallery",
            "Product certificate",
            "News attachment",
            "Sert file",
        ]
        assets = collect_media_library_assets(search_query=search_query, usage_filter=usage_filter)
        context = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Media library",
            "subtitle": "All uploaded and referenced media in one place.",
            "media_assets": assets,
            "search_query": search_query,
            "usage_filter": usage_filter,
            "usage_options": usage_options,
            "asset_count": len(assets),
            **(extra_context or {}),
        }
        return TemplateResponse(request, self.change_list_template, context)

    def delete_asset_view(self, request):
        if request.method != "POST":
            raise PermissionDenied("POST is required to delete media assets.")
        if not self.has_delete_permission(request):
            raise PermissionDenied("You do not have permission to delete media assets.")

        asset_url = (request.POST.get("asset_url") or "").strip() or None
        asset_storage_path = (request.POST.get("asset_storage_path") or "").strip() or None
        asset_title = (request.POST.get("asset_title") or "").strip() or "media asset"
        search_query = (request.POST.get("q") or "").strip()
        usage_filter = (request.POST.get("usage") or "").strip()
        if not asset_url and not asset_storage_path:
            self.message_user(request, "Media asset identifier is missing.", level=messages.ERROR)
            redirect_url = reverse("admin:shop_medialibrary_changelist")
            query_parts = []
            if search_query:
                query_parts.append(f"q={search_query}")
            if usage_filter:
                query_parts.append(f"usage={usage_filter}")
            if query_parts:
                redirect_url = f"{redirect_url}?{'&'.join(query_parts)}"
            return HttpResponseRedirect(redirect_url)

        with transaction.atomic():
            result = delete_media_asset(asset_url, asset_storage_path)
        deleted_dependency_count = sum(result["deleted_rows"].values())
        cleared_field_count = sum(result["cleared_fields"].values()) + result["updated_products"] + result["updated_news"]
        file_message = " File removed from disk." if result["file_deleted"] else " No local file was removed."
        self.message_user(
            request,
            (
                f"Deleted {asset_title}: dependencies removed {deleted_dependency_count}, "
                f"field references cleared {cleared_field_count}, "
                f"products resynced {result['affected_product_count']}."
                f"{file_message}"
            ),
            level=messages.SUCCESS,
        )
        redirect_url = reverse("admin:shop_medialibrary_changelist")
        query_parts = []
        if search_query:
            query_parts.append(f"q={search_query}")
        if usage_filter:
            query_parts.append(f"usage={usage_filter}")
        if query_parts:
            redirect_url = f"{redirect_url}?{'&'.join(query_parts)}"
        return HttpResponseRedirect(redirect_url)


@admin.register(ProductMedia)
class ProductMediaAdmin(admin.ModelAdmin):
    form = ProductMediaAdminForm
    list_display = ("id", "product", "url_preview", "mime_type", "media_kind", "size_bytes", "is_primary", "sort_order")
    list_filter = ("is_primary", "mime_type", "media_kind")
    search_fields = ("product__name", "url", "storage_path")
    readonly_fields = ("url_preview", "storage_path", "url", "mime_type", "media_kind", "size_bytes")
    fields = (
        "product",
        "media_upload",
        "url_preview",
        "storage_path",
        "url",
        "mime_type",
        "media_kind",
        "size_bytes",
        "variants",
        "is_primary",
        "sort_order",
        "alt_text",
    )

    @admin.display(description="Preview")
    def url_preview(self, obj):
        if not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">open</a><br><img src="{0}" style="max-height:120px;max-width:180px;" />', obj.url)

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        ensure_single_primary(obj.product)
        sync_product_media(obj.product)

    def delete_model(self, request, obj):
        product = obj.product
        super().delete_model(request, obj)
        ensure_single_primary(product)
        sync_product_media(product)


@admin.register(ProductGalleryItem)
class ProductGalleryItemAdmin(admin.ModelAdmin):
    form = ProductGalleryItemAdminForm
    list_display = ("id", "product", "title", "file_kind", "mime_type", "size_bytes", "sort_order")
    list_filter = ("file_kind", "mime_type")
    search_fields = ("product__name", "title", "url", "storage_path")
    readonly_fields = ("storage_path", "url", "mime_type", "file_kind", "size_bytes", "document_link")
    fields = (
        "product",
        "gallery_upload",
        "title",
        "document_link",
        "storage_path",
        "url",
        "mime_type",
        "file_kind",
        "size_bytes",
        "sort_order",
    )

    @admin.display(description="Gallery file")
    def document_link(self, obj):
        if not obj or not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">{1}</a>', obj.url, obj.title or "open")


@admin.register(ProductCertificate)
class ProductCertificateAdmin(admin.ModelAdmin):
    form = ProductCertificateAdminForm
    list_display = ("id", "product", "title", "mime_type", "size_bytes", "sort_order")
    search_fields = ("product__name", "title", "url", "storage_path")
    readonly_fields = ("storage_path", "url", "mime_type", "size_bytes", "document_link")
    fields = (
        "product",
        "certificate_upload",
        "title",
        "document_link",
        "storage_path",
        "url",
        "mime_type",
        "size_bytes",
        "sort_order",
    )

    @admin.display(description="Certificate")
    def document_link(self, obj):
        if not obj or not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">{1}</a>', obj.url, obj.title or "open")


@admin.register(NewsAttachment)
class NewsAttachmentAdmin(admin.ModelAdmin):
    form = NewsAttachmentAdminForm
    list_display = ("id", "news", "title", "mime_type", "size_bytes", "sort_order")
    search_fields = ("news__title", "title", "url", "storage_path")
    readonly_fields = ("storage_path", "url", "mime_type", "size_bytes", "document_link")
    fields = (
        "news",
        "attachment_upload",
        "title",
        "document_link",
        "storage_path",
        "url",
        "mime_type",
        "size_bytes",
        "sort_order",
    )

    @admin.display(description="Attachment")
    def document_link(self, obj):
        if not obj or not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">{1}</a>', obj.url, obj.title or "open")


@admin.register(Sert)
class SertAdmin(PublishWorkflowAdminMixin, admin.ModelAdmin):
    form = SertAdminForm
    list_display = ("id", "title", "mime_type", "size_bytes", "sort_order", "status", "updated_at", "updated_by", "preview_link")
    list_filter = ("status", "mime_type")
    search_fields = ("title", "url", "storage_path")
    readonly_fields = ("storage_path", "url", "mime_type", "size_bytes", "document_link", "preview_link", "created_at", "updated_at", "updated_by")
    fields = (
        "title",
        "file_upload",
        "document_link",
        "storage_path",
        "url",
        "mime_type",
        "size_bytes",
        "sort_order",
        "status",
        "preview_link",
        "updated_by",
        "created_at",
        "updated_at",
    )

    @admin.display(description="File")
    def document_link(self, obj):
        if not obj or not obj.url:
            return "No file"
        return format_html('<a href="{0}" target="_blank">{1}</a>', obj.url, obj.title or "open")

    def render_preview_html(self, obj):
        if (obj.mime_type or "").startswith("image/"):
            return f'<section><h1>{escape(obj.title)}</h1><p><img src="{escape(obj.url)}" style="max-width: 100%; height: auto;" /></p></section>'
        if (obj.mime_type or "").startswith("video/"):
            return (
                f'<section><h1>{escape(obj.title)}</h1>'
                f'<video src="{escape(obj.url)}" controls style="max-width: 100%;"></video></section>'
            )
        return (
            f'<section><h1>{escape(obj.title)}</h1>'
            f'<p><a href="{escape(obj.url)}" target="_blank">Open file</a></p></section>'
        )


@admin.register(Characteristic)
class CharacteristicAdmin(admin.ModelAdmin):
    list_display = ("id", "name", "slug", "group", "data_type", "is_filterable", "is_searchable")
    search_fields = ("name", "slug")
    list_filter = ("data_type", "is_filterable", "is_searchable", "group")


@admin.register(ProductCharacteristic)
class ProductCharacteristicAdmin(admin.ModelAdmin):
    form = ProductCharacteristicAdminForm
    list_display = ("id", "product", "characteristic", "value", "created_at")
    search_fields = ("product__name", "characteristic__name", "value")
    list_filter = ("characteristic__group", "characteristic")
    autocomplete_fields = ("product", "characteristic")


admin.site.site_header = "Novotech admin"
admin.site.site_title = "Novotech admin"
admin.site.index_title = "Управление сайтом"


ADMIN_SECTION_ORDER = [
    (
        "catalog",
        "Каталог",
        [
            "shop.Product",
            "shop.Group",
            "shop.Brand",
            "shop.Characteristic",
            "shop.ProductCharacteristic",
        ],
    ),
    (
        "media",
        "Медиа и файлы",
        [
            "shop.MediaLibrary",
            "shop.ProductMedia",
            "shop.ProductGalleryItem",
            "shop.ProductCertificate",
            "shop.Sert",
            "shop.NewsAttachment",
        ],
    ),
    (
        "content",
        "Контент сайта",
        [
            "shop.Slider",
            "shop.News",
            "shop.HtmlContent",
            "shop.ContactInfo",
            "shop.Agent",
            "shop.City",
        ],
    ),
    (
        "orders",
        "Заявки и заказы",
        [
            "shop.Inquiry",
            "shop.PublicOrder",
        ],
    ),
    (
        "settings",
        "Настройки и доступ",
        [
            "shop.OrderEmailSettings",
            "shop.OrderEmailRecipient",
            "auth.User",
            "auth.Group",
        ],
    ),
]

ADMIN_MODEL_NAMES = {
    "shop.Product": "Товары",
    "shop.Group": "Категории",
    "shop.Brand": "Бренды",
    "shop.Characteristic": "Характеристики",
    "shop.ProductCharacteristic": "Значения характеристик",
    "shop.MediaLibrary": "Библиотека медиа",
    "shop.ProductMedia": "Превью товаров",
    "shop.ProductGalleryItem": "Галерея товаров",
    "shop.ProductCertificate": "Сертификаты товаров",
    "shop.Sert": "Общие сертификаты",
    "shop.NewsAttachment": "Файлы новостей",
    "shop.Slider": "Слайдер",
    "shop.News": "Новости",
    "shop.HtmlContent": "HTML-блоки",
    "shop.ContactInfo": "Контакты компании",
    "shop.Agent": "Менеджеры",
    "shop.City": "Города для SEO",
    "shop.Inquiry": "Заявки",
    "shop.PublicOrder": "Заказы",
    "shop.OrderEmailSettings": "Шаблон письма",
    "shop.OrderEmailRecipient": "Получатели писем",
    "auth.User": "Пользователи",
    "auth.Group": "Группы прав",
}


def grouped_admin_app_list(request, app_label=None):
    app_dict = admin.site._build_app_dict(request, app_label)
    model_map = {}
    for current_app_label, app in app_dict.items():
        for model in app.get("models", []):
            key = f"{current_app_label}.{model['object_name']}"
            model = model.copy()
            model["name"] = ADMIN_MODEL_NAMES.get(key, model["name"])
            model_map[key] = model

    grouped_apps = []
    used_keys = set()
    for section_label, section_name, model_keys in ADMIN_SECTION_ORDER:
        models = [model_map[key] for key in model_keys if key in model_map]
        if not models:
            continue
        used_keys.update(key for key in model_keys if key in model_map)
        grouped_apps.append(
            {
                "name": section_name,
                "app_label": section_label,
                "app_url": "",
                "has_module_perms": True,
                "models": models,
            }
        )

    fallback_models = [model for key, model in sorted(model_map.items()) if key not in used_keys]
    if fallback_models:
        grouped_apps.append(
            {
                "name": "Прочее",
                "app_label": "other",
                "app_url": "",
                "has_module_perms": True,
                "models": fallback_models,
            }
        )

    return grouped_apps


admin.site.get_app_list = grouped_admin_app_list
