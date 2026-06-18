import shutil
import tempfile
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from unittest.mock import patch
from zipfile import ZipFile

from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.test.client import RequestFactory
from django.urls import reverse
from openpyxl import load_workbook

from shop.admin_catalog_core_registry import GroupAdmin
from shop.admin import (
    ProductAdmin,
    ProductCertificateAdminForm,
    build_product_export_rows,
    import_products_from_workbook,
)
from shop.models import (
    Brand,
    CatalogImportJob,
    Characteristic,
    Group,
    Product,
    ProductCertificate,
    ProductCharacteristic,
    ProductGalleryItem,
    ProductMedia,
)
from shop.services import catalog_import_jobs as catalog_import_jobs_service
from shop.services.catalog_import import split_title_values
from shop.tests_support import FakeRemoteResponse, TEST_GIF


class GroupAdminChangePageTests(TestCase):
    def setUp(self):
        self.group = Group.objects.create(name="Admin group", slug="admin-group")
        self.brand = Brand.objects.create(name="Admin brand", slug="admin-brand")
        self.product = Product.objects.create(
            sku="GROUP-ADMIN-1",
            name="Admin product",
            price=Decimal("10.00"),
            currency="RUB",
            group=self.group,
            brand=self.brand,
            available=True,
        )
        self.characteristic = Characteristic.objects.create(
            group=self.group,
            name="Admin characteristic",
            slug="admin-characteristic",
            data_type="text",
        )
        ProductCharacteristic.objects.create(
            product=self.product,
            characteristic=self.characteristic,
            value="Admin value",
        )

    def test_group_characteristics_tab_counts_values(self):
        admin = GroupAdmin(Group, AdminSite())

        html = str(admin.characteristics_tab(self.group))

        self.assertIn("Admin characteristic", html)
        self.assertIn("admin-characteristic", html)


class ProductExportAdminTests(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.admin = ProductAdmin(Product, self.site)
        self.factory = RequestFactory()

        self.group = Group.objects.create(name="РўРµРїР»РѕРёР·РѕР»СЏС†РёСЏ", slug="teploizolyatsiya-export")
        self.other_group = Group.objects.create(name="РћРіРЅРµР·Р°С‰РёС‚Р°", slug="ognezashita-export")
        self.brand = Brand.objects.create(name="ENERGOROLL", slug="energoroll-export")
        self.characteristic = Characteristic.objects.create(
            group=self.group,
            name="РўРѕР»С‰РёРЅР° РѕС‚, РјРј",
            slug="tolschina-ot-mm-export",
            data_type="number",
            is_filterable=True,
        )
        self.product = Product.objects.create(
            sku="EXP-0001",
            name="Р­РєСЃРїРѕСЂС‚РёСЂСѓРµРјС‹Р№ С‚РѕРІР°СЂ",
            price=Decimal("150.00"),
            currency="RUB",
            description="РћРїРёСЃР°РЅРёРµ",
            assortment_html="<p>РђСЃСЃРѕСЂС‚РёРјРµРЅС‚ <strong>HTML</strong></p>",
            group=self.group,
            brand=self.brand,
            media=["https://example.com/1.jpg", "https://example.com/2.jpg"],
            available=True,
        )
        self.product.characteristics_html = "<p>РҐР°СЂР°РєС‚РµСЂРёСЃС‚РёРєРё <strong>HTML</strong></p>"
        self.product.save(update_fields=["characteristics_html"])

        ProductCharacteristic.objects.create(
            product=self.product,
            characteristic=self.characteristic,
            value="20",
        )
        Product.objects.create(
            sku="EXP-0002",
            name="РћРіРЅРµР·Р°С‰РёС‚РЅС‹Р№ С‚РѕРІР°СЂ",
            price=Decimal("250.00"),
            currency="RUB",
            group=self.other_group,
            brand=self.brand,
            available=False,
        )

    def test_build_product_export_rows_matches_import_shape(self):
        headers, rows = build_product_export_rows(Product.objects.filter(group=self.group))
        self.assertIn("assortment_html", headers)
        self.assertIn("char_РўРѕР»С‰РёРЅР°_РѕС‚,_РјРј", headers)
        row_map = dict(zip(headers, rows[0]))
        self.assertEqual(row_map["sku"], "EXP-0001")
        self.assertEqual(row_map["group_slug"], "РўРµРїР»РѕРёР·РѕР»СЏС†РёСЏ")
        self.assertEqual(row_map["brand_slug"], "ENERGOROLL")
        self.assertEqual(row_map["assortment_html"], "<p>РђСЃСЃРѕСЂС‚РёРјРµРЅС‚ <strong>HTML</strong></p>")
        self.assertEqual(row_map["media_urls"], "https://example.com/1.jpg,https://example.com/2.jpg")
        self.assertEqual(row_map["char_РўРѕР»С‰РёРЅР°_РѕС‚,_РјРј"], "20")

    def test_export_single_group_workbook_returns_xlsx(self):
        response = self.admin._export_single_group_workbook(self.group)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        values = [cell.value for cell in worksheet[2]]
        row_map = dict(zip(headers, values))
        self.assertEqual(row_map["sku"], "EXP-0001")
        self.assertEqual(row_map["assortment_html"], "<p>РђСЃСЃРѕСЂС‚РёРјРµРЅС‚ <strong>HTML</strong></p>")

    def test_export_grouped_zip_returns_files_per_group(self):
        response = self.admin._export_grouped_zip()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")

        with ZipFile(BytesIO(response.content)) as archive:
            filenames = set(archive.namelist())
            self.assertIn("products_teploizolyatsiya-export.xlsx", filenames)
            self.assertIn("products_ognezashita-export.xlsx", filenames)

    def test_export_includes_characteristics_html_column(self):
        headers, rows = build_product_export_rows(Product.objects.filter(group=self.group))
        self.assertIn("characteristics_html", headers)
        row_map = dict(zip(headers, rows[0]))
        self.assertEqual(row_map["characteristics_html"], self.product.characteristics_html)

    def test_export_includes_new_product_and_attachment_columns(self):
        self.product.search_tsv = "insulation, heat"
        self.product.seo_title = "SEO title"
        self.product.save(update_fields=["search_tsv", "seo_title"])
        ProductGalleryItem.objects.create(
            product=self.product,
            title="Gallery file",
            storage_path="media/gallery.jpg",
            url="/static/gallery.jpg",
            mime_type="image/jpeg",
            file_kind="image",
            size_bytes=10,
        )
        ProductCertificate.objects.create(
            product=self.product,
            title="Certificate",
            storage_path="media/certificate.pdf",
            url="/static/certificate.pdf",
            mime_type="application/pdf",
            size_bytes=10,
        )

        headers, rows = build_product_export_rows(Product.objects.filter(pk=self.product.pk))
        row_map = dict(zip(headers, rows[0]))

        self.assertEqual(row_map["search_tsv"], "insulation, heat")
        self.assertEqual(row_map["seo_title"], "SEO title")
        self.assertEqual(row_map["gallery_urls"], "/static/gallery.jpg")
        self.assertEqual(row_map["certificate_urls"], "/static/certificate.pdf")

    def test_import_uploads_local_file_paths_and_links(self):
        media_root = tempfile.mkdtemp()
        source_dir = tempfile.mkdtemp()
        override = override_settings(MEDIA_ROOT=media_root)
        override.enable()
        try:
            image_path = Path(source_dir) / "preview.jpg"
            image_path.write_bytes(TEST_GIF)
            document_url = "/static/certificate.pdf"
            certificate_path = Path(media_root) / "certificate.pdf"
            certificate_path.write_bytes(TEST_GIF)

            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append([
                "sku",
                "name",
                "price",
                "currency",
                "group_slug",
                "brand_slug",
                "media_urls",
                "certificate_urls",
                "certificate_titles",
                "search_tsv",
                "seo_title",
            ])
            worksheet.append([
                "IMP-LOCAL",
                "Imported local file product",
                "42.50",
                "RUB",
                "Import group",
                "Import brand",
                str(image_path),
                document_url,
                "Certificate link",
                "synonym one",
                "Imported SEO",
            ])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            counters, warnings = import_products_from_workbook(buffer)
            product = Product.objects.get(sku="IMP-LOCAL")
            media = ProductMedia.objects.get(product=product)
            certificate = ProductCertificate.objects.get(product=product)

            self.assertEqual(warnings, [])
            self.assertEqual(counters["media_items_imported"], 1)
            self.assertEqual(counters["certificates_imported"], 1)
            self.assertEqual(product.search_tsv, "synonym one")
            self.assertEqual(product.seo_title, "Imported SEO")
            self.assertTrue(media.url.startswith("/static/admin_uploads/product_media/"))
            self.assertTrue(Path(media.storage_path).exists())
            self.assertEqual(certificate.url, document_url)
            self.assertEqual(certificate.size_bytes, len(TEST_GIF))
            self.assertEqual(Path(certificate.storage_path).read_bytes(), TEST_GIF)
            self.assertEqual(certificate.title, "Certificate link")
        finally:
            override.disable()
            shutil.rmtree(media_root, ignore_errors=True)
            shutil.rmtree(source_dir, ignore_errors=True)

    def test_split_title_values_keeps_commas_inside_title(self):
        self.assertEqual(
            split_title_values(
                "Пароизоляция для кровель, стен и потолка, Сертификат соответствия № 00973/23, "
                "Каркасные конструкции стен, installation g fit"
            ),
            [
                "Пароизоляция для кровель, стен и потолка",
                "Сертификат соответствия № 00973/23",
                "Каркасные конструкции стен",
                "installation g fit",
            ],
        )

    def test_import_strips_char_prefix_from_characteristic_names(self):
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["sku", "name", "price", "currency", "group_slug", "brand_slug", "char_толщина_мкм"])
        worksheet.append(["IMP-CHAR", "Imported characteristic product", "42.50", "RUB", "Import group", "Import brand", "25"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        counters, warnings = import_products_from_workbook(buffer)

        self.assertEqual(warnings, [])
        self.assertEqual(counters["characteristics_created"], 1)
        characteristic = Characteristic.objects.get(group__slug="import-group")
        self.assertEqual(characteristic.name, "толщина мкм")
        self.assertEqual(characteristic.slug, "tolschina-mkm")
        self.assertEqual(
            ProductCharacteristic.objects.get(product__sku="IMP-CHAR", characteristic=characteristic).value,
            "25",
        )

    def test_import_downloads_remote_media_to_local_storage(self):
        media_root = tempfile.mkdtemp()
        override = override_settings(MEDIA_ROOT=media_root)
        override.enable()
        try:
            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["sku", "name", "price", "currency", "media_urls"])
            worksheet.append([
                "IMP-REMOTE",
                "Imported remote media product",
                "55.00",
                "RUB",
                "https://cdn.example.com/preview.jpg",
            ])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            fake_response = FakeRemoteResponse(
                TEST_GIF,
                headers={
                    "Content-Length": str(len(TEST_GIF)),
                    "Content-Disposition": 'attachment; filename="preview.jpg"',
                },
            )
            with patch("shop.services.catalog_import.urlopen", return_value=fake_response):
                counters, warnings = import_products_from_workbook(buffer)

            product = Product.objects.get(sku="IMP-REMOTE")
            media = ProductMedia.objects.get(product=product)

            self.assertEqual(warnings, [])
            self.assertEqual(counters["media_items_imported"], 1)
            self.assertTrue(media.url.startswith("/static/admin_uploads/product_media/"))
            self.assertNotEqual(media.url, "https://cdn.example.com/preview.jpg")
            self.assertTrue(Path(media.storage_path).exists())
            self.assertEqual(Path(media.storage_path).read_bytes(), TEST_GIF)
            self.assertEqual(product.media, [media.url])
        finally:
            override.disable()
            shutil.rmtree(media_root, ignore_errors=True)

    def test_import_downloads_remote_certificate_to_local_storage(self):
        media_root = tempfile.mkdtemp()
        override = override_settings(MEDIA_ROOT=media_root)
        override.enable()
        try:
            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["sku", "name", "price", "currency", "certificate_urls", "certificate_titles"])
            worksheet.append([
                "IMP-REMOTE-CERT",
                "Imported remote certificate product",
                "55.00",
                "RUB",
                "https://cdn.example.com/certificate.pdf",
                "Remote certificate",
            ])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            fake_response = FakeRemoteResponse(
                TEST_GIF,
                headers={
                    "Content-Length": str(len(TEST_GIF)),
                    "Content-Disposition": 'attachment; filename="certificate.pdf"',
                },
            )
            with patch("shop.services.catalog_import.urlopen", return_value=fake_response):
                counters, warnings = import_products_from_workbook(buffer)

            product = Product.objects.get(sku="IMP-REMOTE-CERT")
            certificate = ProductCertificate.objects.get(product=product)

            self.assertEqual(warnings, [])
            self.assertEqual(counters["certificates_imported"], 1)
            self.assertTrue(certificate.url.startswith("/static/admin_uploads/product_certificates/"))
            self.assertNotEqual(certificate.url, "https://cdn.example.com/certificate.pdf")
            self.assertTrue(Path(certificate.storage_path).exists())
            self.assertEqual(Path(certificate.storage_path).read_bytes(), TEST_GIF)
            self.assertEqual(certificate.size_bytes, len(TEST_GIF))
        finally:
            override.disable()
            shutil.rmtree(media_root, ignore_errors=True)

    def test_import_skips_missing_local_certificate_url(self):
        media_root = tempfile.mkdtemp()
        override = override_settings(MEDIA_ROOT=media_root)
        override.enable()
        try:
            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["sku", "name", "price", "currency", "certificate_urls", "certificate_titles"])
            worksheet.append([
                "IMP-MISSING-CERT",
                "Imported missing certificate product",
                "55.00",
                "RUB",
                "/static/admin_uploads/product_certificates/missing.pdf",
                "Missing certificate",
            ])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            counters, warnings = import_products_from_workbook(buffer)

            product = Product.objects.get(sku="IMP-MISSING-CERT")
            self.assertFalse(ProductCertificate.objects.filter(product=product).exists())
            self.assertEqual(counters["certificates_imported"], 0)
            self.assertEqual(len(warnings), 1)
            self.assertEqual(warnings[0]["code"], "certificate_skipped")
            self.assertIn("missing file", warnings[0]["message"])
        finally:
            override.disable()
            shutil.rmtree(media_root, ignore_errors=True)

    def test_localize_remote_media_command_downloads_existing_urls(self):
        media_root = tempfile.mkdtemp()
        override = override_settings(MEDIA_ROOT=media_root)
        override.enable()
        try:
            product = Product.objects.create(
                sku="REMOTE-EXISTING",
                name="Existing remote media product",
                price=Decimal("10.00"),
                currency="RUB",
                media=["https://cdn.example.com/json-preview.jpg"],
                available=True,
            )
            product_media = ProductMedia.objects.create(
                product=product,
                storage_path="",
                url="https://cdn.example.com/preview.jpg",
                mime_type="image/jpeg",
                media_kind="image",
                size_bytes=0,
                is_primary=True,
            )

            def fake_urlopen(request, timeout=30):
                return FakeRemoteResponse(
                    TEST_GIF,
                    headers={
                        "Content-Length": str(len(TEST_GIF)),
                        "Content-Disposition": 'attachment; filename="preview.jpg"',
                    },
                )

            from shop.management.commands import localize_remote_media

            with patch.object(localize_remote_media, "urlopen", side_effect=fake_urlopen):
                call_command("localize_remote_media")

            product.refresh_from_db()
            product_media.refresh_from_db()

            self.assertTrue(product_media.url.startswith("/static/admin_uploads/product_media/"))
            self.assertTrue(Path(product_media.storage_path).exists())
            self.assertTrue(product.media[0].startswith("/static/admin_uploads/products/"))
        finally:
            override.disable()
            shutil.rmtree(media_root, ignore_errors=True)


class AdminFileValidationTests(TestCase):
    def setUp(self):
        self.group = Group.objects.create(name="РўРµСЃС‚", slug="test")
        self.brand = Brand.objects.create(name="Brand", slug="brand")
        self.product = Product.objects.create(
            sku="CERT-1",
            name="РўРѕРІР°СЂ СЃ СЃРµСЂС‚РёС„РёРєР°С‚РѕРј",
            price=Decimal("10.00"),
            currency="RUB",
            group=self.group,
            brand=self.brand,
            available=True,
        )

    def test_new_product_certificate_requires_uploaded_file(self):
        form = ProductCertificateAdminForm(
            data={
                "product": self.product.id,
                "title": "РЎР°Р»Р°Рј Р°Р»РµР№РєСѓРј",
                "sort_order": 0,
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("certificate_upload", form.errors)

    def test_new_product_certificate_with_file_is_valid(self):
        upload = SimpleUploadedFile("certificate.pdf", TEST_GIF, content_type="application/pdf")
        form = ProductCertificateAdminForm(
            data={
                "product": self.product.id,
                "title": "РЎРµСЂС‚РёС„РёРєР°С‚",
                "sort_order": 0,
            },
            files={"certificate_upload": upload},
        )
        self.assertTrue(form.is_valid(), form.errors)


class CatalogImportQueueTests(TestCase):
    def setUp(self):
        self.admin_user = get_user_model().objects.create_superuser(
            username="catalog-import-admin",
            email="catalog-import-admin@example.com",
            password="secret123",
        )
        self.client.force_login(self.admin_user)

    @staticmethod
    def _build_workbook_bytes():
        from openpyxl import Workbook

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["sku", "name", "price", "currency", "group_slug", "brand_slug"])
        worksheet.append(["QUEUE-1", "Queued import product", "99.00", "RUB", "Queued Group", "Queued Brand"])
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def test_admin_import_view_creates_job_and_redirects_to_job_page(self):
        upload = SimpleUploadedFile(
            "products.xlsx",
            self._build_workbook_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with patch("shop.services.import_queue.enqueue_catalog_import_job") as enqueue_mock:
            response = self.client.post(reverse("admin:shop_product_import_xlsx"), {"xlsx_file": upload})

        self.assertEqual(response.status_code, 302)
        job = CatalogImportJob.objects.get()
        self.assertRedirects(response, reverse("admin:shop_catalogimportjob_change", args=[job.pk]))
        self.assertEqual(job.original_filename, "products.xlsx")
        self.assertEqual(job.created_by, self.admin_user)
        enqueue_mock.assert_called_once_with(job.pk)

    def test_process_catalog_import_job_creates_products_in_background(self):
        upload = SimpleUploadedFile(
            "queued-products.xlsx",
            self._build_workbook_bytes(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        with patch("shop.services.import_queue.enqueue_catalog_import_job"):
            job = catalog_import_jobs_service.create_catalog_import_job(upload, created_by=self.admin_user)

        processed_job = catalog_import_jobs_service.process_catalog_import_job(job.pk)

        self.assertEqual(processed_job.status, CatalogImportJob.STATUS_SUCCEEDED)
        self.assertEqual(processed_job.counters["products_created"], 1)
        self.assertTrue(Product.objects.filter(sku="QUEUE-1").exists())
