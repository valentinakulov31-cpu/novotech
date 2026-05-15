import shutil
import tempfile
from io import BytesIO
from zipfile import ZipFile
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

from django.core import mail
from django.contrib.admin.sites import AdminSite
from django.contrib.auth import get_user_model
from django.conf import settings
from django.test import TestCase, override_settings
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test.client import RequestFactory
from django.http import HttpResponse
from rest_framework.test import APIClient
from tinymce.widgets import TinyMCE
from openpyxl import load_workbook

from shop.admin import (
    ContactInfoAdmin,
    HtmlContentAdmin,
    MediaLibraryAdmin,
    HtmlContentAdminForm,
    NewsAdminForm,
    OrderEmailSettingsAdminForm,
    ProductAdmin,
    ProductAdminForm,
    ProductCertificateAdminForm,
    collect_media_library_assets,
    delete_media_asset,
    build_product_export_rows,
    import_products_from_workbook,
    sanitize_catalog_tables,
)
from django_shop.middleware import MediaEmbedHeadersMiddleware
from shop.models import (
    Brand,
    Characteristic,
    City,
    ContactInfo,
    Agent,
    Group,
    HtmlContent,
    Inquiry,
    MediaLibrary,
    News,
    NewsAttachment,
    Product,
    ProductCertificate,
    ProductCharacteristic,
    ProductDocument,
    ProductGalleryItem,
    ProductMedia,
    OrderEmailRecipient,
    OrderEmailSettings,
    PUBLISH_STATUS_DRAFT,
    PUBLISH_STATUS_PUBLISHED,
    PublicOrder,
    Sert,
    Slider,
)
from shop.permissions import IsAdmin


TEST_GIF = (
    b"GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00\xff\xff\xff!"
    b"\xf9\x04\x01\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00"
    b"\x00\x02\x02D\x01\x00;"
)


class FakeRemoteResponse:
    def __init__(self, payload, headers=None):
        self.payload = BytesIO(payload)
        self.headers = headers or {}

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, traceback):
        return False

    def read(self, size=-1):
        return self.payload.read(size)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class CatalogApiTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.client = APIClient()

        self.group = Group.objects.create(name="Теплоизоляция", slug="teploizolyatsiya")
        self.other_group = Group.objects.create(name="Огнезащита", slug="ognezashita")
        self.brand = Brand.objects.create(name="ENERGOROLL", slug="energoroll")
        self.other_brand = Brand.objects.create(name="OBM", slug="obm")

        self.char_thickness = Characteristic.objects.create(
            group=self.group,
            name="Толщина от, мм",
            slug="tolschina-ot-mm",
            data_type="number",
            is_filterable=True,
        )
        self.char_cover = Characteristic.objects.create(
            group=self.group,
            name="Покрытие",
            slug="pokrytie",
            data_type="text",
            is_filterable=True,
        )

        self.product = Product.objects.create(
            sku="ER-0001",
            name="Цилиндры ENERGOROLL RK",
            price=Decimal("100.00"),
            currency="RUB",
            description="Теплоизоляция из минеральной ваты",
            assortment_html="<p><strong>Ассортимент:</strong> цилиндры, маты</p>",
            group=self.group,
            brand=self.brand,
            available=True,
        )
        self.product.characteristics_html = "<div><h3>Характеристики</h3><p>Толщина, покрытие</p></div>"
        self.product.save(update_fields=["characteristics_html"])

        self.other_product = Product.objects.create(
            sku="OB-0001",
            name="Огнезащитный материал ОБМ",
            price=Decimal("250.00"),
            currency="RUB",
            description="Огнезащитный базальтовый материал",
            group=self.other_group,
            brand=self.other_brand,
            available=True,
        )

        ProductCharacteristic.objects.create(product=self.product, characteristic=self.char_thickness, value="20")
        ProductCharacteristic.objects.create(product=self.product, characteristic=self.char_cover, value="Фольга")

        ProductMedia.objects.create(
            product=self.product,
            storage_path="media/a.jpg",
            url="/static/a.jpg",
            mime_type="image/jpeg",
            media_kind="image",
            size_bytes=10,
            is_primary=False,
            sort_order=2,
        )
        ProductMedia.objects.create(
            product=self.product,
            storage_path="media/b.jpg",
            url="/static/b.jpg",
            mime_type="image/jpeg",
            media_kind="image",
            size_bytes=10,
            is_primary=True,
            sort_order=5,
        )
        ProductGalleryItem.objects.create(
            product=self.product,
            title="Видео обзор",
            storage_path="media/video.mp4",
            url="/static/video.mp4",
            mime_type="video/mp4",
            file_kind="video",
            size_bytes=200,
            sort_order=1,
        )
        ProductGalleryItem.objects.create(
            product=self.product,
            title="Фото в интерьере",
            storage_path="media/gallery.jpg",
            url="/static/gallery.jpg",
            mime_type="image/jpeg",
            file_kind="image",
            size_bytes=50,
            sort_order=2,
        )
        ProductDocument.objects.create(
            product=self.product,
            title="Паспорт",
            storage_path="media/passport.pdf",
            url="/static/passport.pdf",
            mime_type="application/pdf",
            size_bytes=120,
            sort_order=1,
        )
        ProductCertificate.objects.create(
            product=self.product,
            title="Сертификат соответствия",
            storage_path="media/certificate.pdf",
            url="/static/certificate.pdf",
            mime_type="application/pdf",
            size_bytes=95,
            sort_order=1,
        )

    def tearDown(self):
        self.override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def test_group_filters_return_facets_with_counts(self):
        response = self.client.get(reverse("group-filters", kwargs={"group_id": self.group.id}))
        self.assertEqual(response.status_code, 200)
        data = response.json()

        self.assertEqual(data["count"], 1)
        self.assertEqual(data["price"]["min"], 100.0)
        self.assertEqual(data["price"]["max"], 100.0)
        self.assertEqual(data["brands"][0]["slug"], "energoroll")

        attr_map = {item["slug"]: item for item in data["attributes"]}
        self.assertIn("tolschina-ot-mm", attr_map)
        self.assertEqual(attr_map["tolschina-ot-mm"]["values"][0]["value"], "20")
        self.assertEqual(attr_map["tolschina-ot-mm"]["values"][0]["count"], 1)
        self.assertEqual(attr_map["tolschina-ot-mm"]["range"]["min"], 20.0)

    def test_products_filter_endpoint_filters_by_query_and_attribute(self):
        response = self.client.post(
            reverse("products-filter"),
            {
                "q": "теплоизоляция ENERGOROLL",
                "group_id": self.group.id,
                "attributes": {
                    "tolschina-ot-mm": ["20"],
                    "pokrytie": ["Фольга"],
                },
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["count"], 1)
        self.assertEqual(data["results"][0]["sku"], "ER-0001")

    def test_products_list_supports_attribute_filters(self):
        response = self.client.get(
            reverse("products-list"),
            {
                "group_id": self.group.id,
                "attr.tolschina-ot-mm": "20",
                "attr.pokrytie": "Фольга",
            },
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["sku"], "ER-0001")

    def test_product_detail_includes_sorted_media_and_documents(self):
        response = self.client.get(reverse("products-detail", kwargs={"product_identifier": self.product.slug}))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["slug"], self.product.slug)
        self.assertEqual([item["url"] for item in data["media_list"]], ["/static/b.jpg", "/static/a.jpg"])
        self.assertEqual([item["url"] for item in data["gallery"]], ["/static/video.mp4", "/static/gallery.jpg"])
        self.assertEqual(data["gallery"][0]["file_kind"], "video")
        self.assertEqual(data["documents_list"][0]["title"], "Паспорт")
        self.assertEqual(data["certificates_list"][0]["title"], "Сертификат соответствия")
        self.assertEqual(data["assortment_html"], "<p><strong>Ассортимент:</strong> цилиндры, маты</p>")

    def test_products_list_includes_media_documents_and_certificates(self):
        response = self.client.get(reverse("products-list"), {"group_id": self.group.id})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data[0]["gallery"][0]["url"], "/static/video.mp4")
        self.assertEqual(data[0]["gallery"][0]["file_kind"], "video")
        self.assertEqual(data[0]["media_list"][0]["media_kind"], "image")
        self.assertEqual(data[0]["documents_list"][0]["url"], "/static/passport.pdf")
        self.assertEqual(data[0]["certificates_list"][0]["url"], "/static/certificate.pdf")
        self.assertEqual(data[0]["assortment_html"], "<p><strong>Ассортимент:</strong> цилиндры, маты</p>")

    def test_public_media_payload_does_not_expose_storage_paths(self):
        response = self.client.get(reverse("products-detail", kwargs={"product_identifier": self.product.slug}))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertNotIn("storage_path", data["media_list"][0])
        self.assertNotIn("storage_path", data["gallery"][0])
        self.assertNotIn("storage_path", data["documents_list"][0])
        self.assertNotIn("storage_path", data["certificates_list"][0])

    def test_products_api_returns_characteristics_html(self):
        detail_response = self.client.get(reverse("products-detail", kwargs={"product_identifier": self.product.slug}))
        self.assertEqual(detail_response.status_code, 200)
        self.assertEqual(
            detail_response.json()["characteristics_html"],
            "<div><h3>Характеристики</h3><p>Толщина, покрытие</p></div>",
        )

        list_response = self.client.get(reverse("products-list"), {"group_id": self.group.id})
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(
            list_response.json()[0]["characteristics_html"],
            "<div><h3>Характеристики</h3><p>Толщина, покрытие</p></div>",
        )

    def test_global_search_uses_characteristics_html(self):
        self.product.characteristics_html = (
            "<table><tr>"
            "<td>Теплопроводность, λ10</td>"
            "<td>Вт/м·°С</td>"
            "<td>0,034</td>"
            "<td>0,034</td>"
            "<td>0,036</td>"
            "<td>ГОСТ 31925-2011</td>"
            "</tr></table>"
        )
        self.product.save(update_fields=["characteristics_html"])

        response = self.client.get(reverse("global-search"), {"q": "Теплопроводность 0,034"})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(any(item["sku"] == "ER-0001" for item in data["results"]["products"]))

    def test_catalog_search_uses_characteristics_html(self):
        self.product.characteristics_html = (
            "<table><tr>"
            "<td>Теплопроводность, λ10</td>"
            "<td>Вт/м·°С</td>"
            "<td>0,034</td>"
            "<td>0,034</td>"
            "<td>0,036</td>"
            "<td>ГОСТ 31925-2011</td>"
            "</tr></table>"
        )
        self.product.save(update_fields=["characteristics_html"])

        response = self.client.post(
            reverse("catalog-results"),
            {
                "context": {"q": "теплопроводность гост"},
                "filters": {},
                "page": 1,
                "page_size": 12,
                "sort": "relevance",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(any(item["sku"] == "ER-0001" for item in data["results"]))

    def test_products_list_supports_popular_random_eight(self):
        for index in range(9):
            Product.objects.create(
                sku=f"POP-{index}",
                name=f"Popular {index}",
                price=Decimal("99.00"),
                currency="RUB",
                group=self.group,
                brand=self.brand,
                available=True,
            )

        response = self.client.get(reverse("products-list"), {"popular": "true"})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data), 8)

    def test_agents_endpoint_returns_published_agents_in_order(self):
        Agent.objects.create(
            full_name="Draft Agent",
            position="Hidden",
            email="draft@example.com",
            phone="+70000000000",
            sort_order=0,
            status=PUBLISH_STATUS_DRAFT,
        )
        Agent.objects.create(
            full_name="Second Agent",
            position="Sales specialist",
            email="second@example.com",
            phone="+79095338586",
            sort_order=2,
            status=PUBLISH_STATUS_PUBLISHED,
        )
        Agent.objects.create(
            full_name="First Agent",
            position="Client manager",
            email="first@example.com",
            phone="+79612283100",
            sort_order=1,
            status=PUBLISH_STATUS_PUBLISHED,
        )

        response = self.client.get(reverse("agents-list"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual([item["full_name"] for item in data], ["First Agent", "Second Agent"])
        self.assertEqual(data[0]["position"], "Client manager")
        self.assertEqual(data[0]["email"], "first@example.com")
        self.assertEqual(data[0]["phone"], "+79612283100")

    def test_product_and_group_endpoints_include_seo_payload(self):
        product_detail = self.client.get(reverse("products-detail", kwargs={"product_identifier": self.product.slug}))
        self.assertEqual(product_detail.status_code, 200)
        product_data = product_detail.json()
        self.assertIn("seo", product_data)
        self.assertEqual(product_data["seo"]["h1"], self.product.name)
        self.assertEqual(product_data["seo"]["canonical_url"], f"/products/{self.product.slug}")

        group_list = self.client.get(reverse("groups-list"))
        self.assertEqual(group_list.status_code, 200)
        groups = group_list.json()
        target_group = next(item for item in groups if item["slug"] == self.group.slug)
        self.assertIn("seo", target_group)
        self.assertEqual(target_group["seo"]["h1"], self.group.name)
        self.assertEqual(target_group["seo"]["canonical_url"], f"/groups/{self.group.slug}")

        group_detail = self.client.get(reverse("groups-detail", kwargs={"group_identifier": self.group.slug}))
        self.assertEqual(group_detail.status_code, 200)
        group_detail_data = group_detail.json()
        self.assertIn("seo", group_detail_data["category"])
        self.assertIn("seo", group_detail_data["products"][0])

    def test_catalog_results_endpoint_uses_shared_context_and_filters(self):
        response = self.client.post(
            reverse("catalog-results"),
            {
                "context": {
                    "q": "Теплоизоляция ENERGOROLL",
                    "group_slug": "teploizolyatsiya",
                    "brand_slug": "energoroll",
                },
                "filters": {
                    "available": True,
                    "attributes": {
                        "pokrytie": ["Фольга"],
                    },
                },
                "page": 1,
                "page_size": 12,
                "sort": "relevance",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["pagination"]["count"], 1)
        self.assertEqual(data["results"][0]["sku"], "ER-0001")
        self.assertEqual(data["context"]["group_slug"], "teploizolyatsiya")
        self.assertEqual(data["context"]["brand_slug"], "energoroll")

    def test_catalog_facets_endpoint_returns_group_brand_and_attribute_facets(self):
        response = self.client.post(
            reverse("catalog-facets"),
            {
                "context": {
                    "q": "ENERGOROLL",
                    "brand_slug": "energoroll",
                },
                "filters": {
                    "attributes": {
                        "pokrytie": ["Фольга"],
                    }
                },
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["summary"]["total_products"], 1)
        self.assertTrue(any(item["slug"] == "energoroll" for item in data["brands"]))
        self.assertTrue(any(item["slug"] == "teploizolyatsiya" for item in data["groups"]))
        attr_map = {item["slug"]: item for item in data["attributes"]}
        self.assertTrue(attr_map["pokrytie"]["values"][0]["selected"])

    def test_catalog_results_endpoint_supports_typo_tolerant_search(self):
        response = self.client.post(
            reverse("catalog-results"),
            {
                "context": {
                    "q": "ENERGORLL",
                },
                "filters": {},
                "page": 1,
                "page_size": 12,
                "sort": "relevance",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(any(item["sku"] == "ER-0001" for item in data["results"]))

    def test_catalog_results_endpoint_uses_search_tsv_synonyms(self):
        response = self.client.post(
            reverse("catalog-results"),
            {
                "context": {
                    "q": "базальтовая изоляция",
                },
                "filters": {},
                "page": 1,
                "page_size": 12,
                "sort": "relevance",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(any(item["sku"] == "ER-0001" for item in data["results"]))


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class GeoSeoApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.city = City.objects.create(
            name="Красноярск",
            slug="krasnoyarsk",
            name_in_prepositional="Красноярске",
            sort_order=1,
            is_active=True,
        )
        self.group = Group.objects.create(name="Теплоизоляция", slug="teploizolyatsiya-geo")
        self.brand = Brand.objects.create(name="ENERGOROLL", slug="energoroll-geo")
        self.product = Product.objects.create(
            sku="GEO-1",
            name="Цилиндры ENERGOROLL GEO",
            price=Decimal("100.00"),
            currency="RUB",
            description="Теплоизоляция для гео-страниц",
            group=self.group,
            brand=self.brand,
            available=True,
        )

    def test_cities_endpoint_returns_active_cities(self):
        response = self.client.get(reverse("cities-list"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data[0]["slug"], "krasnoyarsk")
        self.assertEqual(data[0]["name_in_prepositional"], "Красноярске")

    def test_product_detail_returns_city_aware_seo(self):
        response = self.client.get(
            reverse("products-detail", kwargs={"product_identifier": self.product.slug}),
            {"city_slug": "krasnoyarsk"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["seo"]["city"], "krasnoyarsk")
        self.assertIn("купить в Красноярске", data["seo"]["title"])
        self.assertEqual(data["seo"]["canonical_url"], f"/products/{self.product.slug}")

    def test_product_seo_templates_render_city_placeholders(self):
        self.product.seo_title = "{name} купить {city_prep} | {brand}"
        self.product.seo_description = "{name} {city_prep} со склада"
        self.product.seo_canonical_url = "/geo/{city_slug}/products/{slug}"
        self.product.save(update_fields=["seo_title", "seo_description", "seo_canonical_url"])

        response = self.client.get(
            reverse("products-detail", kwargs={"product_identifier": self.product.slug}),
            {"city_slug": "krasnoyarsk"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["seo"]["title"], "Цилиндры ENERGOROLL GEO купить в Красноярске | ENERGOROLL")
        self.assertEqual(data["seo"]["description"], "Цилиндры ENERGOROLL GEO в Красноярске со склада")
        self.assertEqual(data["seo"]["canonical_url"], f"/geo/krasnoyarsk/products/{self.product.slug}")

    def test_group_detail_returns_city_aware_seo(self):
        response = self.client.get(
            reverse("groups-detail", kwargs={"group_identifier": self.group.slug}),
            {"city_slug": "krasnoyarsk"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["category"]["seo"]["city"], "krasnoyarsk")
        self.assertIn("купить в Красноярске", data["category"]["seo"]["title"])
        self.assertEqual(data["category"]["seo"]["canonical_url"], f"/groups/{self.group.slug}")

    def test_search_returns_city_aware_seo(self):
        response = self.client.get(
            reverse("global-search"),
            {"q": "ENERGOROLL", "city_slug": "krasnoyarsk"},
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["results"]["products"][0]["seo"]["city"], "krasnoyarsk")
        self.assertIn("купить в Красноярске", data["results"]["products"][0]["seo"]["title"])

    def test_catalog_results_return_city_aware_seo(self):
        response = self.client.post(
            reverse("catalog-results"),
            {
                "context": {
                    "group_slug": self.group.slug,
                    "city_slug": "krasnoyarsk",
                },
                "filters": {},
                "page": 1,
                "page_size": 12,
                "sort": "relevance",
            },
            format="json",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["results"][0]["seo"]["city"], "krasnoyarsk")
        self.assertIn("купить в Красноярске", data["results"][0]["seo"]["title"])


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class ProductExportAdminTests(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.admin = ProductAdmin(Product, self.site)
        self.factory = RequestFactory()

        self.group = Group.objects.create(name="Теплоизоляция", slug="teploizolyatsiya-export")
        self.other_group = Group.objects.create(name="Огнезащита", slug="ognezashita-export")
        self.brand = Brand.objects.create(name="ENERGOROLL", slug="energoroll-export")
        self.characteristic = Characteristic.objects.create(
            group=self.group,
            name="Толщина от, мм",
            slug="tolschina-ot-mm-export",
            data_type="number",
            is_filterable=True,
        )
        self.product = Product.objects.create(
            sku="EXP-0001",
            name="Экспортируемый товар",
            price=Decimal("150.00"),
            currency="RUB",
            description="Описание",
            assortment_html="<p>Ассортимент <strong>HTML</strong></p>",
            group=self.group,
            brand=self.brand,
            media=["https://example.com/1.jpg", "https://example.com/2.jpg"],
            available=True,
        )
        self.product.characteristics_html = "<p>Характеристики <strong>HTML</strong></p>"
        self.product.save(update_fields=["characteristics_html"])

        ProductCharacteristic.objects.create(
            product=self.product,
            characteristic=self.characteristic,
            value="20",
        )
        Product.objects.create(
            sku="EXP-0002",
            name="Огнезащитный товар",
            price=Decimal("250.00"),
            currency="RUB",
            group=self.other_group,
            brand=self.brand,
            available=False,
        )

    def test_build_product_export_rows_matches_import_shape(self):
        headers, rows = build_product_export_rows(Product.objects.filter(group=self.group))
        self.assertIn("assortment_html", headers)
        self.assertIn("char_Толщина_от,_мм", headers)
        row_map = dict(zip(headers, rows[0]))
        self.assertEqual(row_map["sku"], "EXP-0001")
        self.assertEqual(row_map["group_slug"], "Теплоизоляция")
        self.assertEqual(row_map["brand_slug"], "ENERGOROLL")
        self.assertEqual(row_map["assortment_html"], "<p>Ассортимент <strong>HTML</strong></p>")
        self.assertEqual(row_map["media_urls"], "https://example.com/1.jpg,https://example.com/2.jpg")
        self.assertEqual(row_map["char_Толщина_от,_мм"], "20")

    def test_export_single_group_workbook_returns_xlsx(self):
        response = self.admin._export_single_group_workbook(self.group)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]
        values = [cell.value for cell in worksheet[2]]
        row_map = dict(zip(headers, values))
        self.assertEqual(row_map["sku"], "EXP-0001")
        self.assertEqual(row_map["assortment_html"], "<p>Ассортимент <strong>HTML</strong></p>")

    def test_export_grouped_zip_returns_files_per_group(self):
        response = self.admin._export_grouped_zip()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")

        with ZipFile(BytesIO(response.content)) as archive:
            filenames = set(archive.namelist())
            self.assertIn("products_teploizolyatsiya-export.xlsx", filenames)
            self.assertIn("products_ognezashita-export.xlsx", filenames)

    def test_export_includes_characteristics_html_column(self):
        headers, rows = build_product_export_rows(Product.objects.filter(group=self.group))
        self.assertIn("characteristics_html", headers)
        row_map = dict(zip(headers, rows[0]))
        self.assertEqual(row_map["characteristics_html"], self.product.characteristics_html)

    def test_export_includes_new_product_and_attachment_columns(self):
        self.product.search_tsv = "insulation, heat"
        self.product.seo_title = "SEO title"
        self.product.save(update_fields=["search_tsv", "seo_title"])
        ProductGalleryItem.objects.create(
            product=self.product,
            title="Gallery file",
            storage_path="media/gallery.jpg",
            url="/static/gallery.jpg",
            mime_type="image/jpeg",
            file_kind="image",
            size_bytes=10,
        )
        ProductDocument.objects.create(
            product=self.product,
            title="Passport",
            storage_path="media/passport.pdf",
            url="/static/passport.pdf",
            mime_type="application/pdf",
            size_bytes=10,
        )
        ProductCertificate.objects.create(
            product=self.product,
            title="Certificate",
            storage_path="media/certificate.pdf",
            url="/static/certificate.pdf",
            mime_type="application/pdf",
            size_bytes=10,
        )

        headers, rows = build_product_export_rows(Product.objects.filter(pk=self.product.pk))
        row_map = dict(zip(headers, rows[0]))

        self.assertEqual(row_map["search_tsv"], "insulation, heat")
        self.assertEqual(row_map["seo_title"], "SEO title")
        self.assertEqual(row_map["gallery_urls"], "/static/gallery.jpg")
        self.assertEqual(row_map["document_titles"], "Passport")
        self.assertEqual(row_map["certificate_urls"], "/static/certificate.pdf")

    def test_import_uploads_local_file_paths_and_links(self):
        media_root = tempfile.mkdtemp()
        source_dir = tempfile.mkdtemp()
        override = override_settings(MEDIA_ROOT=media_root)
        override.enable()
        try:
            image_path = Path(source_dir) / "preview.jpg"
            image_path.write_bytes(TEST_GIF)
            document_url = "https://example.com/passport.pdf"

            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append([
                "sku",
                "name",
                "price",
                "currency",
                "group_slug",
                "brand_slug",
                "media_urls",
                "document_urls",
                "document_titles",
                "search_tsv",
                "seo_title",
            ])
            worksheet.append([
                "IMP-LOCAL",
                "Imported local file product",
                "42.50",
                "RUB",
                "Import group",
                "Import brand",
                str(image_path),
                document_url,
                "Passport link",
                "synonym one",
                "Imported SEO",
            ])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            counters, warnings = import_products_from_workbook(buffer)
            product = Product.objects.get(sku="IMP-LOCAL")
            media = ProductMedia.objects.get(product=product)
            document = ProductDocument.objects.get(product=product)

            self.assertEqual(warnings, [])
            self.assertEqual(counters["media_items_imported"], 1)
            self.assertEqual(counters["documents_imported"], 1)
            self.assertEqual(product.search_tsv, "synonym one")
            self.assertEqual(product.seo_title, "Imported SEO")
            self.assertTrue(media.url.startswith("/static/admin_uploads/product_media/"))
            self.assertTrue(Path(media.storage_path).exists())
            self.assertEqual(document.url, document_url)
            self.assertEqual(document.title, "Passport link")
        finally:
            override.disable()
            shutil.rmtree(media_root, ignore_errors=True)
            shutil.rmtree(source_dir, ignore_errors=True)

    def test_import_downloads_remote_media_to_local_storage(self):
        media_root = tempfile.mkdtemp()
        override = override_settings(MEDIA_ROOT=media_root)
        override.enable()
        try:
            from openpyxl import Workbook

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["sku", "name", "price", "currency", "media_urls"])
            worksheet.append([
                "IMP-REMOTE",
                "Imported remote media product",
                "55.00",
                "RUB",
                "https://cdn.example.com/preview.jpg",
            ])
            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            fake_response = FakeRemoteResponse(
                TEST_GIF,
                headers={
                    "Content-Length": str(len(TEST_GIF)),
                    "Content-Disposition": 'attachment; filename="preview.jpg"',
                },
            )
            with patch("shop.admin.urlopen", return_value=fake_response):
                counters, warnings = import_products_from_workbook(buffer)

            product = Product.objects.get(sku="IMP-REMOTE")
            media = ProductMedia.objects.get(product=product)

            self.assertEqual(warnings, [])
            self.assertEqual(counters["media_items_imported"], 1)
            self.assertTrue(media.url.startswith("/static/admin_uploads/product_media/"))
            self.assertNotEqual(media.url, "https://cdn.example.com/preview.jpg")
            self.assertTrue(Path(media.storage_path).exists())
            self.assertEqual(Path(media.storage_path).read_bytes(), TEST_GIF)
            self.assertEqual(product.media, [media.url])
        finally:
            override.disable()
            shutil.rmtree(media_root, ignore_errors=True)

    def test_localize_remote_media_command_downloads_existing_urls(self):
        media_root = tempfile.mkdtemp()
        override = override_settings(MEDIA_ROOT=media_root)
        override.enable()
        try:
            product = Product.objects.create(
                sku="REMOTE-EXISTING",
                name="Existing remote media product",
                price=Decimal("10.00"),
                currency="RUB",
                media=["https://cdn.example.com/json-preview.jpg"],
                available=True,
            )
            product_media = ProductMedia.objects.create(
                product=product,
                storage_path="",
                url="https://cdn.example.com/preview.jpg",
                mime_type="image/jpeg",
                media_kind="image",
                size_bytes=0,
                is_primary=True,
            )

            def fake_urlopen(request, timeout=30):
                return FakeRemoteResponse(
                    TEST_GIF,
                    headers={
                        "Content-Length": str(len(TEST_GIF)),
                        "Content-Disposition": 'attachment; filename="preview.jpg"',
                    },
                )

            with patch("shop.management.commands.localize_remote_media.urlopen", side_effect=fake_urlopen):
                call_command("localize_remote_media")

            product.refresh_from_db()
            product_media.refresh_from_db()

            self.assertTrue(product_media.url.startswith("/static/admin_uploads/product_media/"))
            self.assertTrue(Path(product_media.storage_path).exists())
            self.assertTrue(product.media[0].startswith("/static/admin_uploads/products/"))
        finally:
            override.disable()
            shutil.rmtree(media_root, ignore_errors=True)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class ProductDocumentUploadTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.client = APIClient()
        self.group = Group.objects.create(name="Тест", slug="test")
        self.brand = Brand.objects.create(name="Brand", slug="brand")
        self.product = Product.objects.create(
            sku="DOC-1",
            name="Товар с документом",
            price=Decimal("10.00"),
            currency="RUB",
            group=self.group,
            brand=self.brand,
            available=True,
        )

        self.original_permission = IsAdmin.has_permission
        IsAdmin.has_permission = lambda self, request, view: True

    def tearDown(self):
        IsAdmin.has_permission = self.original_permission
        self.override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def test_document_upload_endpoint_creates_document(self):
        from django.core.files.uploadedfile import SimpleUploadedFile

        upload = SimpleUploadedFile("passport.pdf", TEST_GIF, content_type="application/pdf")
        response = self.client.post(
            reverse("product-document-upload", kwargs={"product_id": self.product.id}),
            {"file": upload, "title": "Паспорт качества"},
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["title"], "Паспорт качества")
        self.assertTrue(data["url"].startswith("/static/"))
        self.assertEqual(ProductDocument.objects.filter(product=self.product).count(), 1)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class AdminFileValidationTests(TestCase):
    def setUp(self):
        self.group = Group.objects.create(name="Тест", slug="test")
        self.brand = Brand.objects.create(name="Brand", slug="brand")
        self.product = Product.objects.create(
            sku="CERT-1",
            name="Товар с сертификатом",
            price=Decimal("10.00"),
            currency="RUB",
            group=self.group,
            brand=self.brand,
            available=True,
        )

    def test_new_product_certificate_requires_uploaded_file(self):
        form = ProductCertificateAdminForm(
            data={
                "product": self.product.id,
                "title": "Салам алейкум",
                "sort_order": 0,
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("certificate_upload", form.errors)

    def test_new_product_certificate_with_file_is_valid(self):
        upload = SimpleUploadedFile("certificate.pdf", TEST_GIF, content_type="application/pdf")
        form = ProductCertificateAdminForm(
            data={
                "product": self.product.id,
                "title": "Сертификат",
                "sort_order": 0,
            },
            files={"certificate_upload": upload},
        )
        self.assertTrue(form.is_valid(), form.errors)


class MediaEmbedHeadersMiddlewareTests(TestCase):
    def test_allows_iframe_for_target_pdf_paths_only(self):
        factory = RequestFactory()
        middleware = MediaEmbedHeadersMiddleware(lambda request: HttpResponse("ok"))

        allowed_response = middleware(factory.get("/static/admin_uploads/serts/test.pdf"))
        self.assertNotIn("X-Frame-Options", allowed_response)
        self.assertEqual(allowed_response["Cross-Origin-Opener-Policy"], "unsafe-none")
        self.assertEqual(allowed_response["Content-Security-Policy"], "frame-ancestors *")

        blocked_response = middleware(factory.get("/static/admin_uploads/slider/test.jpg"))
        self.assertNotIn("Cross-Origin-Opener-Policy", blocked_response)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class SliderApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        Slider.objects.create(
            image="/static/slider-2.jpg",
            title="Second",
            text="Second text",
            slug="second",
            sort_order=2,
            status=PUBLISH_STATUS_PUBLISHED,
        )
        Slider.objects.create(
            image="/static/slider-1.jpg",
            title="First",
            text="First text",
            slug="first",
            sort_order=1,
            status=PUBLISH_STATUS_PUBLISHED,
        )
        Slider.objects.create(
            image="/static/slider-hidden.jpg",
            title="Hidden",
            text="Hidden text",
            slug="hidden",
            sort_order=0,
            status=PUBLISH_STATUS_DRAFT,
        )

    def test_slider_endpoint_returns_active_items_as_array_in_order(self):
        response = self.client.get(reverse("slider-list"))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual([item["slug"] for item in data], ["first", "second"])
        self.assertNotIn("hidden", [item["slug"] for item in data])


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class NewsAndSertApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.news = News.objects.create(
            title="Новая поставка",
            slug="novaya-postavka",
            content="Приехала новая партия товаров",
            status=PUBLISH_STATUS_PUBLISHED,
        )
        NewsAttachment.objects.create(
            news=self.news,
            title="Каталог PDF",
            storage_path="media/news-catalog.pdf",
            url="/static/news-catalog.pdf",
            mime_type="application/pdf",
            size_bytes=100,
            sort_order=1,
        )
        NewsAttachment.objects.create(
            news=self.news,
            title="Видео обзор",
            storage_path="media/news-video.mp4",
            url="/static/news-video.mp4",
            mime_type="video/mp4",
            size_bytes=500,
            sort_order=2,
        )
        News.objects.create(
            title="Draft news",
            slug="draft-news",
            content="Draft body",
            status=PUBLISH_STATUS_DRAFT,
        )
        Sert.objects.create(
            title="Сертификат ISO",
            storage_path="media/iso.pdf",
            url="/static/iso.pdf",
            mime_type="application/pdf",
            size_bytes=200,
            sort_order=1,
            status=PUBLISH_STATUS_PUBLISHED,
        )
        Sert.objects.create(
            title="Draft sert",
            storage_path="media/draft.pdf",
            url="/static/draft.pdf",
            mime_type="application/pdf",
            size_bytes=50,
            sort_order=2,
            status=PUBLISH_STATUS_DRAFT,
        )

    def test_news_endpoint_returns_attachments(self):
        response = self.client.get(reverse("news-list"))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data), 1)
        self.assertEqual(len(data[0]["attachments"]), 2)
        self.assertEqual(data[0]["attachments"][1]["file_kind"], "video")
        self.assertNotIn("storage_path", data[0]["attachments"][0])

    def test_sert_endpoint_returns_active_files(self):
        response = self.client.get(reverse("sert-list"))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]["url"], "/static/iso.pdf")
        self.assertEqual(data[0]["file_kind"], "document")
        self.assertNotIn("storage_path", data[0])


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class ContentApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        HtmlContent.objects.create(
            title="Draft HTML",
            html_first="<p>Draft first</p>",
            html_second="<p>Draft second</p>",
            status=PUBLISH_STATUS_DRAFT,
        )
        ContactInfo.objects.create(
            title="Draft contacts",
            full_name="Draft manager",
            address="Draft address",
            schedule="Draft schedule",
            phone="+70000000000",
            email="draft@example.com",
            status=PUBLISH_STATUS_DRAFT,
        )
        HtmlContent.objects.create(
            title="Главный HTML",
            html_first="<section><h1>Первый блок</h1></section>",
            html_second="<section><p>Второй блок</p></section>",
            status=PUBLISH_STATUS_PUBLISHED,
        )
        ContactInfo.objects.create(
            title="НАШИ КОНТАКТЫ",
            address="630007, г. Новосибирск,\nпер. Пристанский, 2",
            schedule="ПН-ЧТ: с 9.00 до 18.00\nПТ: с 9.00 до 17.00",
            phone="+7 (383) 263-20-99",
            email="nskteplo-sib.ru",
            status=PUBLISH_STATUS_PUBLISHED,
        )

    def test_html_content_endpoint_returns_two_html_strings(self):
        response = self.client.get(reverse("html-content"))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertNotIn("Draft first", data["html_first"])
        self.assertEqual(data["html_first"], "<section><h1>Первый блок</h1></section>")
        self.assertEqual(data["html_second"], "<section><p>Второй блок</p></section>")

    def test_contact_info_endpoint_returns_contacts(self):
        response = self.client.get(reverse("contact-info"))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertNotIn("id", data)
        self.assertNotIn("updated_at", data)
        self.assertEqual(data["title"], "НАШИ КОНТАКТЫ")
        self.assertEqual(data["phone"], "+7 (383) 263-20-99")
        self.assertIn("Новосибирск", data["address"])


class RichTextAdminFormTests(TestCase):
    def test_html_content_admin_uses_tinymce_widgets(self):
        form = HtmlContentAdminForm()
        self.assertIsInstance(form.fields["html_first"].widget, TinyMCE)
        self.assertIsInstance(form.fields["html_second"].widget, TinyMCE)

    def test_news_admin_uses_tinymce_widget_for_content(self):
        form = NewsAdminForm()
        self.assertIsInstance(form.fields["content"].widget, TinyMCE)

    def test_product_admin_uses_tinymce_widget_for_assortment_html(self):
        form = ProductAdminForm()
        self.assertIsInstance(form.fields["assortment_html"].widget, TinyMCE)

    def test_product_admin_uses_tinymce_widget_for_characteristics_html(self):
        form = ProductAdminForm()
        self.assertIsInstance(form.fields["characteristics_html"].widget, TinyMCE)

    def test_order_email_settings_admin_uses_tinymce_widgets(self):
        form = OrderEmailSettingsAdminForm()
        self.assertIsInstance(form.fields["intro_html"].widget, TinyMCE)
        self.assertIsInstance(form.fields["footer_html"].widget, TinyMCE)

    def test_tinymce_default_config_exposes_catalog_table_style(self):
        config = settings.TINYMCE_DEFAULT_CONFIG
        self.assertIn('/django-static/shop/css/tinymce-content.css', config["content_css"])
        self.assertTrue(any(
            item.get("value") == "var(--catalog-table-border-style)"
            for item in config["table_border_styles"]
        ))
        self.assertEqual(config["table_class_list"], [])
        self.assertEqual(config["style_formats"], [])
        self.assertTrue(config["table_appearance_options"])
        self.assertTrue(config["table_advtab"])
        self.assertNotIn('tableclass', config["toolbar"])

    def test_catalog_table_sanitizer_enforces_catalog_design_inline(self):
        raw_html = (
            '<ul><li>Text<table style="border-style: var(--catalog-table-border-style); border-collapse: collapse; width: 100.033%; '
            'border-width: 1px; height: 44px;" border="1"><tbody>'
            '<tr style="height: 22px;"><td style="width:20%">1</td><td style="width:20%">2</td></tr>'
            '</tbody></table></li></ul>'
        )
        cleaned_html = sanitize_catalog_tables(raw_html)
        self.assertIn('border-style: var(--catalog-table-border-style)', cleaned_html)
        self.assertIn('data-catalog-table="1"', cleaned_html)
        self.assertNotIn('class="catalog-table"', cleaned_html)
        self.assertIn('width: 100%', cleaned_html)
        self.assertIn('background-color: white', cleaned_html)
        self.assertIn('font-size: 14px', cleaned_html)
        self.assertNotIn('border="1"', cleaned_html)
        self.assertIn('padding: 12px', cleaned_html)
        self.assertIn('border: 1px solid #e1e1e1', cleaned_html)
        self.assertIn('text-align: left', cleaned_html)
        self.assertIn('text-align: center', cleaned_html)
        self.assertIn('width: 20%', cleaned_html)

    def test_catalog_table_sanitizer_can_switch_back_to_regular_border_styles(self):
        raw_html = (
            '<table data-catalog-table="1" style="border-style: solid; width: 100%; background-color: white; '
            'border-collapse: collapse; font-size: 14px"><tbody>'
            '<tr><td style="padding: 12px; border: 1px solid #e1e1e1; text-align: left; width: 20%">1</td>'
            '<td style="padding: 12px; border: 1px solid #e1e1e1; text-align: center; width: 20%">2</td></tr>'
            '</tbody></table>'
        )
        cleaned_html = sanitize_catalog_tables(raw_html)
        self.assertNotIn('data-catalog-table="1"', cleaned_html)
        self.assertIn('border-style: solid', cleaned_html)
        self.assertIn('border-width: 1px', cleaned_html)
        self.assertNotIn('background-color: white', cleaned_html)
        self.assertNotIn('font-size: 14px', cleaned_html)
        self.assertNotIn('padding: 12px', cleaned_html)
        self.assertNotIn('text-align: left', cleaned_html)
        self.assertNotIn('text-align: center', cleaned_html)
        self.assertIn('border: 1px solid currentColor', cleaned_html)
        self.assertIn('width: 20%', cleaned_html)

    def test_catalog_table_sanitizer_removes_custom_design_when_marker_is_gone(self):
        raw_html = (
            '<table data-catalog-table="1" style="width: 100%; background-color: white; '
            'border-collapse: collapse; font-size: 14px; border-width: 1px"><tbody>'
            '<tr><td style="padding: 12px; border: 1px solid #e1e1e1; text-align: left; width: 20%">1</td>'
            '<td style="padding: 12px; border: 1px solid #e1e1e1; text-align: center; width: 20%">2</td></tr>'
            '</tbody></table>'
        )
        cleaned_html = sanitize_catalog_tables(raw_html)
        self.assertNotIn('data-catalog-table="1"', cleaned_html)
        self.assertNotIn('background-color: white', cleaned_html)
        self.assertIn('border-collapse: collapse', cleaned_html)
        self.assertNotIn('font-size: 14px', cleaned_html)
        self.assertNotIn('padding: 12px', cleaned_html)
        self.assertNotIn('border: 1px solid #e1e1e1', cleaned_html)
        self.assertNotIn('text-align: left', cleaned_html)
        self.assertNotIn('text-align: center', cleaned_html)
        self.assertIn('border-width: 1px', cleaned_html)
        self.assertIn('width: 20%', cleaned_html)

    def test_signature_only_table_is_not_forced_into_catalog_design(self):
        raw_html = (
            '<table style="width: 100%; background-color: white; border-collapse: collapse; '
            'font-size: 14px; border-width: 1px" border="1"><tbody>'
            '<tr><td>1</td><td>2</td></tr></tbody></table>'
        )
        cleaned_html = sanitize_catalog_tables(raw_html)
        self.assertEqual(cleaned_html, raw_html)

    def test_regular_border_styles_are_pushed_to_cells_for_frontend_rendering(self):
        raw_html = (
            '<table style="border-collapse: collapse; width: 100.033%; border-width: 1px; border-style: dotted;" border="1">'
            '<tbody><tr><td>1</td><td>2</td></tr></tbody></table>'
        )
        cleaned_html = sanitize_catalog_tables(raw_html)
        self.assertIn('border-style: dotted', cleaned_html)
        self.assertIn('border-width: 1px', cleaned_html)
        self.assertIn('border-collapse: collapse', cleaned_html)
        self.assertIn('border: 1px dotted currentColor', cleaned_html)
        self.assertNotIn('background-color: white', cleaned_html)
        self.assertNotIn('padding: 12px', cleaned_html)
        self.assertNotIn('border="1"', cleaned_html)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class InquiryApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()

    def test_inquiry_endpoint_creates_record(self):
        response = self.client.post(
            reverse("inquiry-create"),
            {
                "name": "Иван",
                "phone": "+79990001122",
                "email": "ivan@example.com",
                "message": "Нужна консультация по товару",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data["name"], "Иван")
        self.assertEqual(Inquiry.objects.count(), 1)
        inquiry = Inquiry.objects.first()
        self.assertEqual(inquiry.phone, "+79990001122")

    def test_inquiry_endpoint_allows_missing_phone_and_email(self):
        response = self.client.post(
            reverse("inquiry-create"),
            {
                "name": "Петр",
                "message": "Перезвоните мне завтра",
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201)
        inquiry = Inquiry.objects.get(name="Петр")
        self.assertIsNone(inquiry.phone)
        self.assertIsNone(inquiry.email)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class PublicOrderApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.group = Group.objects.create(name="Теплоизоляция", slug="teploizolyatsiya")
        self.brand = Brand.objects.create(name="ENERGOROLL", slug="energoroll")
        self.product_one = Product.objects.create(
            sku="ER-0001",
            name="Цилиндры ENERGOROLL RK",
            price=Decimal("100.00"),
            currency="RUB",
            group=self.group,
            brand=self.brand,
            available=True,
        )
        self.product_two = Product.objects.create(
            sku="ER-0002",
            name="Маты ENERGOROLL",
            price=Decimal("200.00"),
            currency="RUB",
            group=self.group,
            brand=self.brand,
            available=True,
        )

    def test_public_order_endpoint_creates_order_with_items(self):
        response = self.client.post(
            reverse("public-order-create"),
            {
                "name": "Иван",
                "phone": "+79990001122",
                "email": "ivan@example.com",
                "address": "Красноярск, ул. Тестовая, 1",
                "comment": "Позвоните перед доставкой",
                "items": [
                    {"product_id": self.product_one.id, "qty": 2},
                    {"product_id": self.product_two.id, "qty": 1},
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 201)
        data = response.json()
        self.assertEqual(data["name"], "Иван")
        self.assertEqual(data["total_items"], 3)
        self.assertEqual(len(data["items"]), 2)
        self.assertEqual(PublicOrder.objects.count(), 1)
        order = PublicOrder.objects.get()
        self.assertEqual(order.address, "Красноярск, ул. Тестовая, 1")
        self.assertEqual(order.items.count(), 2)

    def test_public_order_requires_phone_and_items(self):
        response = self.client.post(
            reverse("public-order-create"),
            {
                "name": "Петр",
                "items": [],
            },
            format="json",
        )
        self.assertEqual(response.status_code, 400)
        data = response.json()
        self.assertIn("phone", data)


@override_settings(
    ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"],
    EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
)
class PublicOrderEmailTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        mail.outbox = []
        self.group = Group.objects.create(name="Теплоизоляция", slug="teploizolyatsiya-email")
        self.brand = Brand.objects.create(name="ENERGOROLL", slug="energoroll-email")
        self.product = Product.objects.create(
            sku="EMAIL-1",
            name="Цилиндры ENERGOROLL RK",
            price=Decimal("100.00"),
            currency="RUB",
            group=self.group,
            brand=self.brand,
            available=True,
        )
        self.recipient = OrderEmailRecipient.objects.create(email="sales@example.com", name="Sales", is_active=True)
        self.email_settings = OrderEmailSettings.objects.create(
            title="Order notifications",
            subject="Заказ с фронта",
            intro_html="<p>Новый заказ на сайте.</p>",
            footer_html="<p>Проверьте заявку в админке.</p>",
            from_email="robot@example.com",
            status=PUBLISH_STATUS_PUBLISHED,
        )

    def test_public_order_sends_email_to_active_recipients(self):
        response = self.client.post(
            reverse("public-order-create"),
            {
                "name": "Иван",
                "phone": "+79990001122",
                "email": "ivan@example.com",
                "address": "Красноярск, ул. Тестовая, 1",
                "comment": "Позвоните перед доставкой",
                "items": [{"product_id": self.product.id, "qty": 2}],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].to, ["sales@example.com"])
        self.assertEqual(mail.outbox[0].subject, "Заказ с фронта")
        self.assertIn("EMAIL-1", mail.outbox[0].body)
        self.assertIn("Иван", mail.outbox[0].body)

    def test_public_order_skips_email_without_active_recipients(self):
        self.recipient.is_active = False
        self.recipient.save(update_fields=["is_active"])

        response = self.client.post(
            reverse("public-order-create"),
            {
                "name": "Петр",
                "phone": "+79990001111",
                "items": [{"product_id": self.product.id, "qty": 1}],
            },
            format="json",
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(len(mail.outbox), 0)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class GlobalSearchFuzzyTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.group = Group.objects.create(name="Теплоизоляция", slug="teploizolyatsiya-fuzzy")
        self.brand = Brand.objects.create(name="ENERGOROLL", slug="energoroll-fuzzy")
        self.product = Product.objects.create(
            sku="FUZZY-1",
            name="Цилиндры ENERGOROLL RK",
            price=Decimal("150.00"),
            currency="RUB",
            search_tsv="минвата, базальтовая изоляция, энерго ролл",
            group=self.group,
            brand=self.brand,
            available=True,
        )
        self.cover_characteristic = Characteristic.objects.create(
            group=self.group,
            name="Покрытие",
            slug="pokrytie-fuzzy",
            data_type="text",
            is_filterable=True,
        )
        ProductCharacteristic.objects.create(
            product=self.product,
            characteristic=self.cover_characteristic,
            value="Без покрытия",
        )

    def test_global_search_supports_typos(self):
        response = self.client.get(reverse("global-search"), {"q": "ENERGORLL"})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(any(item["sku"] == "FUZZY-1" for item in data["results"]["products"]))
        self.assertTrue(any(item["slug"] == "energoroll-fuzzy" for item in data["results"]["brands"]))

    def test_global_search_uses_search_tsv_synonyms(self):
        self.product.search_tsv = "энерго ролл, цилиндры rk, базальтовая изоляция"
        self.product.save(update_fields=["search_tsv"])

        response = self.client.get(reverse("global-search"), {"q": "базальтовая изоляция"})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(any(item["sku"] == "FUZZY-1" for item in data["results"]["products"]))

    def test_global_search_does_not_duplicate_products_for_characteristic_matches(self):
        response = self.client.get(reverse("global-search"), {"q": "покрыт"})
        self.assertEqual(response.status_code, 200)
        data = response.json()
        product_ids = [item["id"] for item in data["results"]["products"]]
        self.assertEqual(len(product_ids), len(set(product_ids)))


class IsAdminPermissionTests(TestCase):
    def test_staff_user_passes_permission(self):
        request = RequestFactory().get("/admin-only")
        request.user = get_user_model().objects.create_user(
            username="staff-user",
            password="secret123",
            is_staff=True,
        )

        self.assertTrue(IsAdmin().has_permission(request, view=None))

    def test_regular_user_is_rejected(self):
        request = RequestFactory().get("/admin-only")
        request.user = get_user_model().objects.create_user(
            username="regular-user",
            password="secret123",
        )

        self.assertFalse(IsAdmin().has_permission(request, view=None))


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class AdminPublishWorkflowTests(TestCase):
    def setUp(self):
        self.admin_user = get_user_model().objects.create_superuser(
            username="workflow-admin",
            email="workflow-admin@example.com",
            password="secret123",
        )
        self.site = AdminSite()

    def test_html_content_preview_is_available_for_draft_content(self):
        content = HtmlContent.objects.create(
            title="Draft HTML preview",
            html_first="<p>Draft first</p>",
            html_second="<p>Draft second</p>",
            status=PUBLISH_STATUS_DRAFT,
        )
        admin_instance = HtmlContentAdmin(HtmlContent, self.site)
        request = RequestFactory().get(reverse("admin:shop_htmlcontent_preview", args=[content.pk]))
        request.user = self.admin_user

        response = admin_instance.preview_view(request, str(content.pk))

        self.assertEqual(response.status_code, 200)
        rendered = response.rendered_content
        self.assertIn("Draft HTML preview", rendered)
        self.assertIn("Draft first", rendered)

    def test_contact_info_preview_is_available_for_draft_content(self):
        contact = ContactInfo.objects.create(
            title="Draft contacts",
            full_name="Draft manager",
            address="Draft address",
            schedule="Draft schedule",
            phone="+70000000000",
            email="draft@example.com",
            status=PUBLISH_STATUS_DRAFT,
        )
        admin_instance = ContactInfoAdmin(ContactInfo, self.site)
        request = RequestFactory().get(reverse("admin:shop_contactinfo_preview", args=[contact.pk]))
        request.user = self.admin_user

        response = admin_instance.preview_view(request, str(contact.pk))

        self.assertEqual(response.status_code, 200)
        rendered = response.rendered_content
        self.assertIn("Draft contacts", rendered)
        self.assertIn("Draft manager", rendered)


@override_settings(ALLOWED_HOSTS=["testserver", "localhost", "127.0.0.1"])
class MediaLibraryAdminTests(TestCase):
    def setUp(self):
        self.media_root = tempfile.mkdtemp()
        self.override = override_settings(MEDIA_ROOT=self.media_root)
        self.override.enable()
        self.site = AdminSite()
        self.admin = MediaLibraryAdmin(MediaLibrary, self.site)
        self.product = Product.objects.create(
            sku="MEDIA-1",
            name="Media product",
            price=Decimal("100.00"),
            currency="RUB",
            available=True,
        )
        self.brand = Brand.objects.create(name="Media brand", slug="media-brand", media="/static/admin_uploads/brands/shared.jpg")
        self.news = News.objects.create(title="Media news", slug="media-news", content="Body", media=["/static/admin_uploads/brands/shared.jpg"])
        self.shared_storage_path = Path(self.media_root) / "admin_uploads" / "brands" / "shared.jpg"
        self.shared_storage_path.parent.mkdir(parents=True, exist_ok=True)
        self.shared_storage_path.write_bytes(TEST_GIF)
        ProductMedia.objects.create(
            product=self.product,
            storage_path=str(self.shared_storage_path),
            url="/static/admin_uploads/brands/shared.jpg",
            mime_type="image/jpeg",
            media_kind="image",
            size_bytes=len(TEST_GIF),
            is_primary=True,
            sort_order=0,
        )

    def tearDown(self):
        self.override.disable()
        shutil.rmtree(self.media_root, ignore_errors=True)

    def test_collect_media_library_assets_groups_shared_usages(self):
        assets = collect_media_library_assets()
        shared_asset = next(item for item in assets if item["url"] == "/static/admin_uploads/brands/shared.jpg")
        self.assertEqual(shared_asset["usage_count"], 3)
        self.assertEqual(shared_asset["kind"], "image")
        self.assertTrue(shared_asset["file_exists"])
        usage_sources = {usage["source_label"] for usage in shared_asset["usages"]}
        self.assertIn("Brand image", usage_sources)
        self.assertIn("News media field", usage_sources)
        self.assertIn("Product media", usage_sources)

    def test_delete_media_asset_removes_file_and_all_dependencies(self):
        result = delete_media_asset(
            asset_url="/static/admin_uploads/brands/shared.jpg",
            asset_storage_path=str(self.shared_storage_path),
        )

        self.brand.refresh_from_db()
        self.news.refresh_from_db()
        self.product.refresh_from_db()

        self.assertEqual(self.brand.media, None)
        self.assertIsNone(self.news.media)
        self.assertIsNone(self.product.media)
        self.assertEqual(ProductMedia.objects.count(), 0)
        self.assertFalse(self.shared_storage_path.exists())
        self.assertEqual(result["deleted_rows"]["product_media"], 1)
        self.assertEqual(result["cleared_fields"]["brands"], 1)
        self.assertEqual(result["updated_news"], 1)
        self.assertTrue(result["file_deleted"])

    def test_media_library_admin_changelist_renders_custom_table(self):
        request = RequestFactory().get("/admin/shop/medialibrary/")
        request.user = get_user_model().objects.create_superuser(
            username="media-admin",
            email="media-admin@example.com",
            password="secret123",
        )

        response = self.admin.changelist_view(request)

        self.assertEqual(response.status_code, 200)
        self.assertIn("Media library", response.rendered_content)
        self.assertIn("Delete everywhere", response.rendered_content)
