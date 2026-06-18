import mimetypes
import re
import shutil
import time
import uuid
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from urllib.parse import unquote, urlparse
from urllib.request import Request, urlopen

from django.conf import settings
from django.core.exceptions import ValidationError
from django.db import transaction

from shop.file_utils import infer_file_kind
from shop.model_utils import transliterate_slug
from shop.models import (
    Brand,
    Characteristic,
    Group,
    Product,
    ProductCertificate,
    ProductCharacteristic,
    ProductGalleryItem,
    ProductMedia,
    SharedProductGallery,
)
from shop.services.media_assets import ensure_single_primary, sync_product_media


REMOTE_IMPORT_MAX_BYTES = 100 * 1024 * 1024
REMOTE_IMPORT_TIMEOUT_SECONDS = 5
REMOTE_IMPORT_TOTAL_TIMEOUT_SECONDS = 20
PRODUCT_OPTIONAL_IMPORT_FIELDS = (
    "description",
    "assortment_html",
    "characteristics_html",
    "search_tsv",
    "seo_title",
    "seo_h1",
    "seo_description",
    "seo_keywords",
    "seo_canonical_url",
    "seo_robots",
)


def build_import_issue(
    level: str,
    row_number: int | None,
    sku: str,
    message: str,
    *,
    column: str = "",
    value=None,
    code: str = "",
) -> dict:
    return {
        "level": level,
        "row_number": row_number,
        "sku": sku,
        "column": column,
        "value": "" if value in (None, "") else str(value),
        "code": code,
        "message": message,
    }


def parse_bool(value):
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().lower()
    if normalized in {"да", "true", "1", "yes", "y"}:
        return True
    if normalized in {"нет", "false", "0", "no", "n"}:
        return False
    return bool(normalized)


def parse_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    text = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValidationError(f"Не удалось распознать число: {value}") from exc


def split_media_urls(value):
    if not value:
        return None
    if isinstance(value, list):
        urls = [str(item).strip() for item in value if str(item).strip()]
        return urls or None
    urls = [item.strip() for item in re.split(r"[\n;,]+", str(value)) if item.strip()]
    return urls or None


def split_title_values(value):
    items = split_media_urls(value)
    return items or []


def is_probable_url(value: str | None) -> bool:
    if not value:
        return False
    parsed = urlparse(str(value).strip())
    return parsed.scheme in {"http", "https"} or str(value).startswith(str(settings.MEDIA_URL))


def normalize_local_file_path(value: str | None) -> Path | None:
    text = str(value or "").strip().strip('"')
    if not text:
        return None
    parsed = urlparse(text)
    if parsed.scheme == "file":
        text = parsed.path
        if re.match(r"^/[a-zA-Z]:/", text):
            text = text[1:]
    if is_probable_url(text):
        return None
    path = Path(text).expanduser()
    if path.exists() and path.is_file():
        return path
    return None


def save_local_file_path(source_path: Path, folder_name: str) -> dict:
    media_root = Path(settings.MEDIA_ROOT)
    target_dir = media_root / "admin_uploads" / folder_name
    target_dir.mkdir(parents=True, exist_ok=True)

    if source_path.stat().st_size <= 0:
        raise ValidationError(f"Imported file is empty: {source_path}")

    storage_path = target_dir / f"{uuid.uuid4().hex}{source_path.suffix}"
    shutil.copyfile(source_path, storage_path)

    relative_path = storage_path.relative_to(media_root).as_posix()
    mime_type = mimetypes.guess_type(source_path.name)[0] or "application/octet-stream"
    return {
        "storage_path": str(storage_path),
        "url": f"{settings.MEDIA_URL}{relative_path}",
        "mime_type": mime_type,
        "size_bytes": storage_path.stat().st_size,
        "title": source_path.name,
    }


def filename_from_remote_response(url: str, response) -> str:
    content_disposition = response.headers.get("Content-Disposition", "")
    match = re.search(r'filename\*?=(?:UTF-8\'\')?"?([^";]+)"?', content_disposition, flags=re.IGNORECASE)
    if match:
        filename = unquote(match.group(1)).strip()
        if filename:
            return Path(filename).name
    parsed_name = Path(unquote(urlparse(url).path)).name
    return parsed_name or f"remote-{uuid.uuid4().hex[:8]}"


def save_remote_file_url(url: str, folder_name: str) -> dict:
    media_root = Path(settings.MEDIA_ROOT)
    target_dir = media_root / "admin_uploads" / folder_name
    target_dir.mkdir(parents=True, exist_ok=True)

    request = Request(url, headers={"User-Agent": "Mozilla/5.0 (compatible; NovatehMediaImporter/1.0)"})
    started_at = time.monotonic()
    with urlopen(request, timeout=REMOTE_IMPORT_TIMEOUT_SECONDS) as response:
        content_length = response.headers.get("Content-Length")
        if content_length and int(content_length) > REMOTE_IMPORT_MAX_BYTES:
            raise ValidationError(f"Удалённый файл больше {REMOTE_IMPORT_MAX_BYTES // 1024 // 1024} МБ: {url}")

        original_name = filename_from_remote_response(url, response)
        storage_path = target_dir / f"{uuid.uuid4().hex}{Path(original_name).suffix}"
        downloaded = 0

        with storage_path.open("wb") as destination:
            while True:
                if time.monotonic() - started_at > REMOTE_IMPORT_TOTAL_TIMEOUT_SECONDS:
                    storage_path.unlink(missing_ok=True)
                    raise ValidationError(f"Истекло время ожидания скачивания удалённого файла: {url}")
                chunk = response.read(1024 * 1024)
                if not chunk:
                    break
                downloaded += len(chunk)
                if downloaded > REMOTE_IMPORT_MAX_BYTES:
                    storage_path.unlink(missing_ok=True)
                    raise ValidationError(f"Удалённый файл больше {REMOTE_IMPORT_MAX_BYTES // 1024 // 1024} МБ: {url}")
                destination.write(chunk)

    if storage_path.stat().st_size <= 0:
        storage_path.unlink(missing_ok=True)
        raise ValidationError(f"Remote file is empty: {url}")

    relative_path = storage_path.relative_to(media_root).as_posix()
    mime_type = mimetypes.guess_type(original_name)[0] or mimetypes.guess_type(str(storage_path))[0] or "application/octet-stream"
    return {
        "storage_path": str(storage_path),
        "url": f"{settings.MEDIA_URL}{relative_path}",
        "mime_type": mime_type,
        "size_bytes": storage_path.stat().st_size,
        "title": original_name,
    }


def resolve_import_file_reference(value, folder_name: str):
    text = str(value or "").strip()
    if not text:
        return None, None

    local_path = normalize_local_file_path(text)
    if local_path:
        return save_local_file_path(local_path, folder_name), None

    parsed = urlparse(text)
    if parsed.scheme in {"http", "https"}:
        try:
            return save_remote_file_url(text, folder_name), None
        except Exception as exc:  # noqa: BLE001
            return None, f"не удалось скачать удалённый файл: {text}. {exc}"

    if is_probable_url(text):
        mime_type = mimetypes.guess_type(urlparse(text).path)[0] or "application/octet-stream"
        if text.startswith(str(settings.MEDIA_URL)):
            relative_path = text[len(str(settings.MEDIA_URL)):].lstrip("/")
            candidate_path = Path(settings.MEDIA_ROOT) / relative_path
            if not candidate_path.exists() or not candidate_path.is_file():
                return None, f"local media URL points to a missing file: {text}"
            size_bytes = candidate_path.stat().st_size
            if size_bytes <= 0:
                return None, f"local media URL points to an empty file: {text}"
            storage_path = str(candidate_path)
        else:
            return None, f"unsupported file URL: {text}"
        return {
            "storage_path": storage_path,
            "url": text,
            "mime_type": mime_type,
            "size_bytes": size_bytes,
            "title": Path(urlparse(text).path).name or text,
        }, None

    return None, f"не найден путь или URL к файлу: {text}"


def infer_data_type(value) -> str:
    if value in (None, ""):
        return "text"
    if isinstance(value, (int, float, Decimal)):
        return "number"
    text = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        Decimal(text)
        return "number"
    except InvalidOperation:
        return "text"


def characteristic_header_from_name(name: str) -> str:
    normalized = str(name or "").strip().replace("\n", " ")
    normalized = re.sub(r"\s+", "_", normalized)
    return f"char_{normalized}"


def characteristic_name_from_header(header: str) -> str:
    normalized = str(header or "").strip().replace("\n", " ")
    if normalized.lower().startswith("char_"):
        normalized = normalized[5:]
    normalized = normalized.replace("_", " ")
    return re.sub(r"\s+", " ", normalized).strip()


def clean_optional_text(value):
    text = str(value or "").strip()
    return text or None


def resolve_group(raw_value: str):
    if not raw_value:
        return None, False
    value = str(raw_value).strip()
    slug = transliterate_slug(value)
    group = Group.objects.filter(slug=slug).first() or Group.objects.filter(name=value).first()
    if group:
        return group, False
    return Group.objects.create(name=value, slug=slug), True


def resolve_brand(raw_value: str):
    if not raw_value:
        return None, False
    value = str(raw_value).strip()
    slug = transliterate_slug(value)
    brand = Brand.objects.filter(slug=slug).first() or Brand.objects.filter(name=value).first()
    if brand:
        return brand, False
    return Brand.objects.create(name=value, slug=slug), True


def resolve_shared_gallery(raw_value: str):
    if not raw_value:
        return None, False
    value = str(raw_value).strip()
    slug = transliterate_slug(value)
    gallery = SharedProductGallery.objects.filter(slug=slug).first() or SharedProductGallery.objects.filter(name=value).first()
    if gallery:
        return gallery, False
    return SharedProductGallery.objects.create(name=value, slug=slug), True


def resolve_characteristic(group: Group, header: str, sample_value):
    name = characteristic_name_from_header(header)
    slug = transliterate_slug(name)
    characteristic = Characteristic.objects.filter(group=group, slug=slug).first()
    if characteristic:
        updated = False
        inferred_type = infer_data_type(sample_value)
        if not characteristic.name:
            characteristic.name = name
            updated = True
        if not characteristic.data_type:
            characteristic.data_type = inferred_type
            updated = True
        if updated:
            characteristic.save(update_fields=["name", "data_type"])
        return characteristic, False
    return Characteristic.objects.create(
        group=group,
        name=name,
        slug=slug,
        data_type=infer_data_type(sample_value),
        is_filterable=True,
        is_searchable=False,
    ), True


def sync_imported_product_media(product, values, issues, row_number):
    ProductMedia.objects.filter(product=product).delete()
    created_items = []
    for index, value in enumerate(values):
        uploaded, warning = resolve_import_file_reference(value, "product_media")
        if warning:
            issues.append(
                build_import_issue(
                    "warning",
                    row_number,
                    product.sku,
                    f"Медиа пропущено: {warning}",
                    column="media_urls",
                    value=value,
                    code="media_skipped",
                )
            )
            continue
        created_items.append(
            ProductMedia.objects.create(
                product=product,
                storage_path=uploaded["storage_path"],
                url=uploaded["url"],
                mime_type=uploaded["mime_type"],
                media_kind=infer_file_kind(uploaded["mime_type"]),
                size_bytes=uploaded["size_bytes"],
                is_primary=index == 0,
                sort_order=index,
                alt_text=product.name,
            )
        )
    ensure_single_primary(product)
    sync_product_media(product)
    return len(created_items)


def sync_imported_product_gallery(product, values, title_values, issues, row_number):
    ProductGalleryItem.objects.filter(product=product).delete()
    created_count = 0
    for index, value in enumerate(values):
        uploaded, warning = resolve_import_file_reference(value, "product_gallery")
        if warning:
            issues.append(
                build_import_issue(
                    "warning",
                    row_number,
                    product.sku,
                    f"Файл галереи пропущен: {warning}",
                    column="gallery_urls",
                    value=value,
                    code="gallery_skipped",
                )
            )
            continue
        ProductGalleryItem.objects.create(
            product=product,
            title=title_values[index] if index < len(title_values) else uploaded["title"],
            storage_path=uploaded["storage_path"],
            url=uploaded["url"],
            mime_type=uploaded["mime_type"],
            file_kind=infer_file_kind(uploaded["mime_type"]),
            size_bytes=uploaded["size_bytes"],
            sort_order=index,
        )
        created_count += 1
    return created_count


def sync_imported_titled_files(model, product, values, title_values, folder_name, issues, row_number, warning_label):
    model.objects.filter(product=product).delete()
    created_count = 0
    for index, value in enumerate(values):
        uploaded, warning = resolve_import_file_reference(value, folder_name)
        if warning:
            issues.append(
                build_import_issue(
                    "warning",
                    row_number,
                    product.sku,
                    f"{warning_label.capitalize()} пропущен: {warning}",
                    column=f"{warning_label}_urls",
                    value=value,
                    code=f"{warning_label}_skipped",
                )
            )
            continue
        model.objects.create(
            product=product,
            title=title_values[index] if index < len(title_values) else uploaded["title"],
            storage_path=uploaded["storage_path"],
            url=uploaded["url"],
            mime_type=uploaded["mime_type"],
            size_bytes=uploaded["size_bytes"],
            sort_order=index,
        )
        created_count += 1
    return created_count


def import_products_from_workbook(workbook_file):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValidationError("openpyxl не установлен. Добавьте пакет в окружение перед импортом XLSX.") from exc

    workbook = load_workbook(workbook_file, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration as exc:
        raise ValidationError("XLSX-файл пустой.") from exc

    required_headers = {"sku", "name", "price", "currency"}
    missing = sorted(required_headers - set(headers))
    if missing:
        raise ValidationError(f"Не хватает обязательных колонок: {', '.join(missing)}")

    char_headers = [header for header in headers if header.startswith("char_")]
    counters = {
        "products_created": 0,
        "products_updated": 0,
        "groups_created": 0,
        "brands_created": 0,
        "shared_galleries_created": 0,
        "characteristics_created": 0,
        "product_characteristics_upserted": 0,
        "media_items_imported": 0,
        "gallery_items_imported": 0,
        "certificates_imported": 0,
        "rows_skipped": 0,
    }
    issues = []

    with transaction.atomic():
        for row_number, row in enumerate(rows, start=2):
            payload = dict(zip(headers, row))
            sku = str(payload.get("sku") or "").strip()
            if not sku:
                counters["rows_skipped"] += 1
                continue

            name = str(payload.get("name") or "").strip()
            if not name:
                issues.append(
                    build_import_issue(
                        "warning",
                        row_number,
                        sku,
                        "Строка пропущена: пустое название товара.",
                        column="name",
                        code="empty_name",
                    )
                )
                counters["rows_skipped"] += 1
                continue

            group_name = str(payload.get("group_slug") or "").strip()
            brand_name = str(payload.get("brand_slug") or "").strip()
            shared_gallery_name = str(payload.get("shared_gallery_slug") or "").strip()
            group, group_created = resolve_group(group_name) if group_name else (None, False)
            brand, brand_created = resolve_brand(brand_name) if brand_name else (None, False)
            shared_gallery, shared_gallery_created = resolve_shared_gallery(shared_gallery_name) if shared_gallery_name else (None, False)
            counters["groups_created"] += int(group_created)
            counters["brands_created"] += int(brand_created)
            counters["shared_galleries_created"] += int(shared_gallery_created)

            defaults = {
                "name": name,
                "price": parse_decimal(payload.get("price")),
                "currency": str(payload.get("currency") or "RUB").strip() or "RUB",
                "group": group,
                "brand": brand,
                "shared_gallery": shared_gallery,
                "available": parse_bool(payload.get("available")),
            }
            for field_name in PRODUCT_OPTIONAL_IMPORT_FIELDS:
                if field_name in payload:
                    defaults[field_name] = clean_optional_text(payload.get(field_name))
            if "slug" in payload:
                defaults["slug"] = clean_optional_text(payload.get("slug")) or ""
            if "media_urls" in payload:
                defaults["media"] = split_media_urls(payload.get("media_urls"))

            product, created = Product.objects.update_or_create(sku=sku, defaults=defaults)
            counters["products_created"] += int(created)
            counters["products_updated"] += int(not created)

            if "media_urls" in payload:
                counters["media_items_imported"] += sync_imported_product_media(
                    product,
                    split_media_urls(payload.get("media_urls")) or [],
                    issues,
                    row_number,
                )
            if "gallery_urls" in payload:
                counters["gallery_items_imported"] += sync_imported_product_gallery(
                    product,
                    split_media_urls(payload.get("gallery_urls")) or [],
                    split_title_values(payload.get("gallery_titles")),
                    issues,
                    row_number,
                )
            if "certificate_urls" in payload:
                counters["certificates_imported"] += sync_imported_titled_files(
                    ProductCertificate,
                    product,
                    split_media_urls(payload.get("certificate_urls")) or [],
                    split_title_values(payload.get("certificate_titles")),
                    "product_certificates",
                    issues,
                    row_number,
                    "certificate",
                )

            if not group and char_headers:
                has_characteristics = any(payload.get(header) not in (None, "") for header in char_headers)
                if has_characteristics:
                    issues.append(
                        build_import_issue(
                            "warning",
                            row_number,
                            sku,
                            "Характеристики пропущены: не указана категория.",
                            column="group_slug",
                            code="missing_group_for_characteristics",
                        )
                    )
                continue

            for header in char_headers:
                raw_value = payload.get(header)
                if raw_value in (None, ""):
                    continue
                characteristic, char_created = resolve_characteristic(group, header, raw_value)
                counters["characteristics_created"] += int(char_created)
                ProductCharacteristic.objects.update_or_create(
                    product=product,
                    characteristic=characteristic,
                    defaults={"value": str(raw_value).strip()},
                )
                counters["product_characteristics_upserted"] += 1

    return counters, issues


def build_product_export_rows(products_queryset):
    products = list(
        products_queryset.select_related("group", "brand", "shared_gallery").prefetch_related(
            "characteristics__characteristic",
            "media_files",
            "gallery_items",
            "certificates",
        )
    )
    characteristic_names = {}
    for product in products:
        for product_characteristic in product.characteristics.all():
            characteristic_names[product_characteristic.characteristic.slug] = product_characteristic.characteristic.name

    ordered_characteristics = sorted(characteristic_names.items(), key=lambda item: item[1].lower())
    char_headers = [characteristic_header_from_name(name) for _, name in ordered_characteristics]

    headers = [
        "sku",
        "slug",
        "name",
        "price",
        "currency",
        "description",
        "assortment_html",
        "characteristics_html",
        "search_tsv",
        "seo_title",
        "seo_h1",
        "seo_description",
        "seo_keywords",
        "seo_canonical_url",
        "seo_robots",
        "group_slug",
        "brand_slug",
        "shared_gallery_slug",
        "available",
        "media_urls",
        "gallery_urls",
        "gallery_titles",
        "certificate_urls",
        "certificate_titles",
        *char_headers,
    ]
    rows = []
    for product in products:
        values_by_slug = {
            product_characteristic.characteristic.slug: product_characteristic.value
            for product_characteristic in product.characteristics.all()
        }
        product_media_urls = [item.url for item in product.media_files.all()]
        rows.append(
            [
                product.sku,
                product.slug,
                product.name,
                str(product.price),
                product.currency,
                product.description or "",
                product.assortment_html or "",
                product.characteristics_html or "",
                product.search_tsv or "",
                product.seo_title or "",
                product.seo_h1 or "",
                product.seo_description or "",
                product.seo_keywords or "",
                product.seo_canonical_url or "",
                product.seo_robots or "",
                product.group.name if product.group else "",
                product.brand.name if product.brand else "",
                product.shared_gallery.slug if product.shared_gallery else "",
                "Да" if product.available else "Нет",
                ",".join(product_media_urls or (product.media or [])),
                ",".join(item.url for item in product.gallery_items.all()),
                ",".join(item.title or "" for item in product.gallery_items.all()),
                ",".join(item.url for item in product.certificates.all()),
                ",".join(item.title or "" for item in product.certificates.all()),
                *[values_by_slug.get(characteristic_slug, "") or "" for characteristic_slug, _ in ordered_characteristics],
            ]
        )
    return headers, rows


def workbook_bytes_from_headers_and_rows(headers, rows, title="Products"):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ValidationError("openpyxl не установлен. Добавьте пакет в окружение перед экспортом XLSX.") from exc

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title[:31] or "Products"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def workbook_bytes_from_import_report(counters, issues):
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ValidationError("openpyxl не установлен. Добавьте пакет в окружение перед экспортом XLSX.") from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Показатель", "Значение"])
    for key, value in counters.items():
        summary.append([key, value])

    details = workbook.create_sheet("Issues")
    details.append(["Уровень", "Строка", "SKU", "Колонка", "Код", "Значение", "Сообщение"])
    for issue in issues:
        details.append(
            [
                issue.get("level", ""),
                issue.get("row_number", ""),
                issue.get("sku", ""),
                issue.get("column", ""),
                issue.get("code", ""),
                issue.get("value", ""),
                issue.get("message", ""),
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
